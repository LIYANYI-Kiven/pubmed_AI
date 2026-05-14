#!/usr/bin/env python3
"""
PubMed 植物糖基化文献下载工具（精简版，配置从 config.json 读取）
"""

import json
import os
import time
from typing import Dict, List
import xml.etree.ElementTree as ET

import requests
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


def load_config(config_path: str = "config.json") -> Dict[str, any]:
    """
    从 JSON 文件加载配置。
    query 不需要提前转义，由 requests 自动处理 URL encoding。
    """
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"配置文件不存在: {config_path}")
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


class PubMedConfig:
    """
    从 config.json 读取的 PubMed 配置。
    """
    def __init__(self, d: Dict[str, any]):
        self.date_start = d["date_start"]
        self.date_end = d["date_end"]
        self.api_key = d.get("api_key", None)
        self.batch_size = int(d.get("batch_size", 100))
        self.sleep_sec = float(d.get("sleep_sec", 0.4))
        self.out_file = d["out_file"]  # 修正这里的 typo
        self.query = d["query"]  # query 本身不转义，留给 requests 处理


class PubMedClient:
    """
    PubMed E‑utils 客户端，封装 esearch / efetch。
    """
    BASE_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"

    def __init__(self, api_key: str = None):
        self.params = {"api_key": api_key} if api_key else {}

    def get_count(
        self,
        query: str,
        date_start: str,
        date_end: str,
        timeout: float = 30.0,
    ) -> int:
        """
        获取符合条件的 PubMed 文献总数。
        """
        r = requests.get(
            self.BASE_URL + "esearch.fcgi",
            params={
                "db": "pubmed",
                "term": query,
                "datetype": "pdat",
                "mindate": date_start,
                "maxdate": date_end,
                "retmode": "json",
                "retmax": 0,
                **self.params,
            },
            timeout=timeout,
        )
        r.raise_for_status()
        return int(r.json()["esearchresult"]["count"])

    def get_pmids(
        self,
        query: str,
        date_start: str,
        date_end: str,
        batch_size: int,
        start: int,
        timeout: float = 30.0,
    ) -> List[str]:
        """
        从指定位置获取一批 PMID。
        """
        r = requests.get(
            self.BASE_URL + "esearch.fcgi",
            params={
                "db": "pubmed",
                "term": query,
                "datetype": "pdat",
                "mindate": date_start,
                "maxdate": date_end,
                "retmode": "json",
                "retmax": min(batch_size, 1000),
                "retstart": start,
                **self.params,
            },
            timeout=timeout,
        )
        r.raise_for_status()
        return r.json()["esearchresult"].get("idlist", [])

    def fetch_xml(
        self,
        pmids: List[str],
        timeout: float = 60.0,
    ) -> ET.Element:
        """
        从 PubMed 拉取一组文献的 XML。
        """
        r = requests.get(
            self.BASE_URL + "efetch.fcgi",
            params={
                "db": "pubmed",
                "id": ",".join(pmids),
                "retmode": "xml",
                "rettype": "abstract",
                **self.params,
            },
            timeout=timeout,
        )
        r.raise_for_status()
        return ET.fromstring(r.content)


def parse_article(article: ET.Element) -> Dict[str, str]:
    """
    从一个 PubMed <PubmedArticle> 解析出基本信息。
    """
    # PMID
    pmid = article.findtext(".//PMID", "").strip()

    # Title
    title_elem = article.find(".//ArticleTitle")
    title = "".join(title_elem.itertext()) if title_elem is not None else ""

    # Abstract
    abstract = " ".join(
        (f"[{a.attrib.get('Label', '')}] " if a.attrib.get("Label") else "")
        + (a.text or "")
        for a in article.findall(".//AbstractText")
        if a.text
    ).strip()

    # First Author
    first_author = ""
    author = article.find(".//Author")
    if author is not None:
        last = author.findtext("LastName", "").strip()
        fore = author.findtext("ForeName", "").strip()
        first_author = f"{last} {fore}".strip()

    # MeSH Terms
    mesh_terms = [
        m.text for m in article.findall(".//MeshHeading/DescriptorName") if m.text
    ]
    mesh_str = "; ".join(mesh_terms)

    # Publication Date
    year = article.findtext(".//ArticleDate/Year", "").strip()
    if year:
        month = article.findtext(".//ArticleDate/Month", "").strip().zfill(2)
        day = article.findtext(".//ArticleDate/Day", "").strip().zfill(2)
        pub_date = f"{year}-{month}-{day}"
    else:
        year = article.findtext(".//PubDate/Year", "").strip()
        month = article.findtext(".//PubDate/Month", "").strip()
        day = article.findtext(".//PubDate/Day", "").strip()
        if year and month and day:
            pub_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        elif year and month:
            pub_date = f"{year}-{month}"
        elif year:
            pub_date = year
        else:
            pub_date = article.findtext(".//PubDate/MedlineDate", "").strip()

    return {
        "PMID": pmid,
        "Title": title,
        "Abstract": abstract,
        "First Author": first_author,
        "MeSH Terms": mesh_str,
        "Publication Date": pub_date,
    }


def fetch_records(config: PubMedConfig, verbose: bool = True) -> List[Dict[str, str]]:
    """
    从 PubMed 获取所有匹配记录列表。
    """
    print("\n[1/3] 获取 PMID 列表...")
    cli = PubMedClient(config.api_key)

    total = cli.get_count(config.query, config.date_start, config.date_end)
    print(f"  命中：{total} 篇")

    pmids = []
    for start in range(0, total, config.batch_size):
        batch = cli.get_pmids(
            query=config.query,
            date_start=config.date_start,
            date_end=config.date_end,
            batch_size=config.batch_size,
            start=start,
        )
        pmids.extend(batch)
        if verbose:
            print(f"  进度：{len(pmids)}/{total}", end="\r", flush=True)
        time.sleep(config.sleep_sec)
    print(f"\n  完成：获取 {len(pmids)} 个 PMID")

    print("\n[2/3] 下载文章信息...")
    records = []
    for i in range(0, len(pmids), config.batch_size):
        chunk = pmids[i : i + config.batch_size]

        try:
            root = cli.fetch_xml(chunk)
        except (requests.RequestException, ET.ParseError) as e:
            print(f"\n  警告：批次 {i // config.batch_size + 1} 下载/解析失败: {e}")
            time.sleep(1)
            continue

        for article in root.findall(".//PubmedArticle"):
            records.append(parse_article(article))

        done = min(i + config.batch_size, len(pmids))
        if verbose:
            print(
                f"  进度：{done}/{len(pmids)} ({100 * done // len(pmids):d}%)  已解析：{len(records)} 篇",
                end="\r",
                flush=True,
            )
        time.sleep(config.sleep_sec)
    print(f"\n  完成：解析 {len(records)} 篇文章")

    return records


def export_excel(records: List[Dict[str, str]], out_file: str, verbose: bool = True):
    """
    导出到带格式的 Excel。
    """
    df = pd.DataFrame(records)
    df["_sort"] = pd.to_datetime(df["Publication Date"], errors="coerce")
    df = df.sort_values("_sort", ascending=False).drop(columns=["_sort"])
    df.reset_index(drop=True, inplace=True)
    df.index += 1

    os.makedirs(os.path.dirname(out_file), exist_ok=True)

    if verbose:
        print("\n[3/3] 导出 Excel...")

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=True, index_label="No.", sheet_name="Results")
        ws = writer.sheets["Results"]

        # 表头样式
        fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 冻结首行 + 自动筛选
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

        # 列宽
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 80
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 50
        ws.column_dimensions["G"].width = 14

        # 隔行变色
        even_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = even_fill if row_idx % 2 == 0 else odd_fill
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row_idx].height = 50

    if verbose:
        print(f"\n{'='*70}")
        print(f"  ✅ 完成！")
        print(f"  📁 文件：{out_file}")
        print(f"  📊 总计：{len(df)} 篇文章")
        print(f"  🔍 检索词：{records[0].get('Title', '')[:80]}...")
        print(f"{'='*70}\n")


def main():
    """
    主入口。
    """
    print("=" * 70)
    print("  PubMed 植物糖基化文献下载工具（精简版）")
    print("=" * 70)

    # 读配置
    cfg = PubMedConfig(load_config("config.json"))
    print(f"  时间范围：{cfg.date_start} ~ {cfg.date_end}")
    print(f"  检索词：{cfg.query[:80]}...")

    # 获取记录
    records = fetch_records(cfg, verbose=True)

    # 导出 Excel
    export_excel(records, cfg.out_file, verbose=True)


if __name__ == "__main__":
    main()