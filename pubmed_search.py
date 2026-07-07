#!/usr/bin/env python3
"""
PubMed 文献检索工具（增强版）

在 pubmed-improved.py 的基础上增加：
  1. 同时输出 Excel 和 JSON 两种格式
  2. JSON 字段更丰富：关键词、DOI、PMCID、期刊、所有作者、完整链接等
  3. JSON 结构参考 LitSense，每条记录包含 pmid / pmcid / links / metadata 等

配置文件: config.json（与 pubmed-improved.py 完全兼容）

用法:
    python pubmed_search.py
    python pubmed_search.py --config config.json
"""

import argparse
import json
import os
import re
import time
from typing import Dict, List, Optional
import xml.etree.ElementTree as ET

import requests
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


# ---------------------------------------------------------------------------
# 配置
# ---------------------------------------------------------------------------

def load_config(path: str = "config.json") -> Dict:
    if not os.path.exists(path):
        raise FileNotFoundError(f"配置文件不存在: {path}")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


# ---------------------------------------------------------------------------
# PubMed E-utils 客户端（含 PMCID / DOI 获取）
# ---------------------------------------------------------------------------

class PubMedClient:
    BASE_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"

    def __init__(self, api_key: Optional[str] = None):
        self.params = {"api_key": api_key} if api_key else {}
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": "pubmed-search/2.0"})

    def search_with_history(self, query: str, date_start: str,
                            date_end: str, timeout: float = 30.0) -> Dict:
        r = self.session.get(
            self.BASE_URL + "esearch.fcgi",
            params={
                "db": "pubmed", "term": query,
                "datetype": "pdat", "mindate": date_start, "maxdate": date_end,
                "retmode": "json", "retmax": 0,
                "usehistory": "y", "sort": "pub_date",
                **self.params,
            },
            timeout=timeout,
        )
        r.raise_for_status()
        result = r.json()["esearchresult"]
        return {
            "count": int(result["count"]),
            "webenv": result.get("webenv"),
            "query_key": result.get("querykey"),
        }

    def get_pmids_with_history(self, webenv: str, query_key: str,
                               batch_size: int, start: int,
                               timeout: float = 30.0) -> List[str]:
        r = self.session.get(
            self.BASE_URL + "esearch.fcgi",
            params={
                "db": "pubmed", "retmode": "json",
                "retmax": min(batch_size, 10000), "retstart": start,
                "webenv": webenv, "query_key": query_key,
                **self.params,
            },
            timeout=timeout,
        )
        r.raise_for_status()
        return r.json()["esearchresult"].get("idlist", [])

    def fetch_xml(self, pmids: List[str], timeout: float = 90.0) -> ET.Element:
        r = self.session.get(
            self.BASE_URL + "efetch.fcgi",
            params={
                "db": "pubmed", "id": ",".join(pmids),
                "retmode": "xml", "rettype": "abstract",
                **self.params,
            },
            timeout=timeout,
        )
        r.raise_for_status()
        return ET.fromstring(r.content)

    def fetch_pmcids(self, pmids: List[str], timeout: float = 30.0) -> Dict[str, str]:
        """
        通过 elink 批量获取 PMID → PMCID 映射。
        返回 {pmid: pmcid_or_empty}
        """
        if not pmids:
            return {}
        try:
            r = self.session.get(
                self.BASE_URL + "elink.fcgi",
                params={
                    "dbfrom": "pubmed", "db": "pmc",
                    "id": ",".join(pmids), "retmode": "json",
                    **self.params,
                },
                timeout=timeout,
            )
            r.raise_for_status()
            data = r.json()
            mapping = {}
            for linkset in data.get("linksets", []):
                src_ids = linkset.get("ids", [])
                pmc_links = []
                for ls_db in linkset.get("linksetdbs", []):
                    if ls_db.get("dbto") == "pmc":
                        pmc_links = ls_db.get("links", [])
                        break
                for pmid in src_ids:
                    mapping[str(pmid)] = f"PMC{pmc_links[0]}" if pmc_links else ""
            return mapping
        except Exception:
            return {}


# ---------------------------------------------------------------------------
# 文章解析（扩展版，提取更多字段）
# ---------------------------------------------------------------------------

def _text(elem: Optional[ET.Element]) -> str:
    if elem is None:
        return ""
    return "".join(elem.itertext()).strip()


def parse_article_full(article: ET.Element) -> Dict:
    """
    解析 PubmedArticle XML，返回包含所有字段的字典。
    包含 xlsx 所需字段 + JSON 额外字段。
    """
    medline = article.find(".//MedlineCitation")
    article_node = article.find(".//Article")

    # ---- 基本 ----
    pmid = article.findtext(".//PMID", "").strip()

    # ---- 标题 ----
    title_elem = article_node.find("ArticleTitle") if article_node is not None else None
    title = _text(title_elem)

    # ---- 摘要 ----
    abstract_parts = []
    if article_node is not None:
        for a in article_node.findall(".//AbstractText"):
            label = a.attrib.get("Label", "")
            text  = a.text or ""
            if label:
                abstract_parts.append(f"[{label}] {text}")
            else:
                abstract_parts.append(text)
    abstract = " ".join(p for p in abstract_parts if p).strip()

    # ---- 作者 ----
    authors = []
    for au in article.findall(".//Author"):
        last  = (au.findtext("LastName") or "").strip()
        fore  = (au.findtext("ForeName") or "").strip()
        collective = (au.findtext("CollectiveName") or "").strip()
        if last:
            authors.append(f"{last} {fore}".strip())
        elif collective:
            authors.append(collective)
    first_author = authors[0] if authors else ""
    all_authors  = "; ".join(authors)

    # ---- 期刊 ----
    journal_elem  = article.find(".//Journal")
    journal_title = ""
    journal_iso   = ""
    issn          = ""
    volume = issue = pages = ""
    if journal_elem is not None:
        journal_title = (journal_elem.findtext("Title") or "").strip()
        journal_iso   = (journal_elem.findtext("ISOAbbreviation") or "").strip()
        issn          = (journal_elem.findtext("ISSN") or "").strip()
        ji = journal_elem.find("JournalIssue")
        if ji is not None:
            volume = (ji.findtext("Volume") or "").strip()
            issue  = (ji.findtext("Issue") or "").strip()
    pages = (article.findtext(".//Pagination/MedlinePgn") or "").strip()

    # ---- 发表日期 ----
    pub_date = ""
    year  = (article.findtext(".//ArticleDate/Year")  or "").strip()
    month = (article.findtext(".//ArticleDate/Month") or "").strip()
    day   = (article.findtext(".//ArticleDate/Day")   or "").strip()
    if year:
        pub_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}" if month and day else year
    else:
        year  = (article.findtext(".//PubDate/Year")  or "").strip()
        month = (article.findtext(".//PubDate/Month") or "").strip()
        day   = (article.findtext(".//PubDate/Day")   or "").strip()
        if year and month and day:
            pub_date = f"{year}-{month.zfill(2)}-{day.zfill(2)}"
        elif year and month:
            pub_date = f"{year}-{month}"
        elif year:
            pub_date = year
        else:
            pub_date = (article.findtext(".//PubDate/MedlineDate") or "").strip()

    # ---- DOI ----
    doi = ""
    for eid in article.findall(".//ArticleId"):
        if eid.attrib.get("IdType") == "doi":
            doi = (eid.text or "").strip()
            break

    # ---- PMC ID ----
    pmcid = ""
    for eid in article.findall(".//ArticleId"):
        if eid.attrib.get("IdType") == "pmc":
            pmcid = (eid.text or "").strip()
            break

    # ---- 关键词 ----
    keywords = [kw.text.strip() for kw in article.findall(".//Keyword") if kw.text]

    # ---- MeSH Terms ----
    mesh_terms = [
        m.text for m in article.findall(".//MeshHeading/DescriptorName") if m.text
    ]

    # ---- 发表类型 ----
    pub_types = [
        pt.text for pt in article.findall(".//PublicationType") if pt.text
    ]

    # ---- 语言 ----
    language = (article.findtext(".//Language") or "").strip()

    # ---- 链接 ----
    pubmed_url = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else ""
    pmc_url    = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmcid}/" if pmcid else ""
    doi_url    = f"https://doi.org/{doi}" if doi else ""
    fulltext_url = pmc_url or doi_url  # 优先 PMC（开放获取），其次 DOI

    return {
        # xlsx 列（与 pubmed-improved.py 完全兼容）
        "PMID":             pmid,
        "Title":            title,
        "Abstract":         abstract,
        "First Author":     first_author,
        "MeSH Terms":       "; ".join(mesh_terms),
        "Publication Date": pub_date,
        # 额外列（xlsx 新增）
        "All Authors":      all_authors,
        "Journal":          journal_title,
        "Journal Abbr":     journal_iso,
        "ISSN":             issn,
        "Volume":           volume,
        "Issue":            issue,
        "Pages":            pages,
        "DOI":              doi,
        "PMCID":            pmcid,
        "Keywords":         "; ".join(keywords),
        "Publication Types": "; ".join(pub_types),
        "Language":         language,
        # 链接
        "PubMed URL":       pubmed_url,
        "PMC URL":          pmc_url,
        "DOI URL":          doi_url,
        "Full Text URL":    fulltext_url,
    }


def record_to_json_entry(record: Dict) -> Dict:
    """
    将解析后的记录转换为 JSON 格式条目。
    结构参考 LitSense 输出风格，同时包含完整元数据。
    """
    pmid  = record.get("PMID", "")
    pmcid = record.get("PMCID", "")
    doi   = record.get("DOI", "")

    return {
        "pmid":             pmid,
        "pmcid":            pmcid or None,
        "doi":              doi or None,
        "title":            record.get("Title", ""),
        "abstract":         record.get("Abstract", ""),
        "authors": {
            "first_author": record.get("First Author", ""),
            "all_authors":  record.get("All Authors", ""),
        },
        "journal": {
            "title":    record.get("Journal", ""),
            "abbr":     record.get("Journal Abbr", ""),
            "issn":     record.get("ISSN", ""),
            "volume":   record.get("Volume", ""),
            "issue":    record.get("Issue", ""),
            "pages":    record.get("Pages", ""),
        },
        "publication_date": record.get("Publication Date", ""),
        "publication_types": [
            pt.strip() for pt in record.get("Publication Types", "").split(";") if pt.strip()
        ],
        "language":         record.get("Language", ""),
        "keywords": [
            kw.strip() for kw in record.get("Keywords", "").split(";") if kw.strip()
        ],
        "mesh_terms": [
            m.strip() for m in record.get("MeSH Terms", "").split(";") if m.strip()
        ],
        "links": {
            "pubmed":    record.get("PubMed URL", ""),
            "pmc":       record.get("PMC URL", "") or None,
            "doi":       record.get("DOI URL", "") or None,
            "full_text": record.get("Full Text URL", "") or None,
        },
    }


# ---------------------------------------------------------------------------
# 检索主流程
# ---------------------------------------------------------------------------

def fetch_records(query: str, date_start: str, date_end: str,
                  api_key: Optional[str], batch_size: int,
                  sleep_sec: float) -> List[Dict]:
    """从 PubMed 检索并返回完整解析的记录列表。"""
    cli = PubMedClient(api_key)

    print("\n[1/3] 初始化搜索...")
    search = cli.search_with_history(query, date_start, date_end)
    total     = search["count"]
    webenv    = search["webenv"]
    query_key = search["query_key"]
    print(f"  ✅ 命中：{total} 篇")

    print("\n[2/3] 获取 PMID 列表...")
    pmids = []
    for start in range(0, total, batch_size):
        try:
            batch = cli.get_pmids_with_history(webenv, query_key, batch_size, start)
            pmids.extend(batch)
            print(f"  进度：{len(pmids)}/{total} ({100*len(pmids)//max(total,1)}%)",
                  end="\r", flush=True)
            time.sleep(sleep_sec)
        except Exception as e:
            print(f"\n  ⚠️  批次 {start} 失败: {e}")
            time.sleep(1)
    print(f"\n  ✅ 获取 {len(pmids)} 个 PMID")

    print("\n[3/3] 下载文章详情...")
    records = []
    for i in range(0, len(pmids), batch_size):
        chunk = pmids[i:i + batch_size]
        try:
            root = cli.fetch_xml(chunk)
            for article in root.findall(".//PubmedArticle"):
                records.append(parse_article_full(article))
        except Exception as e:
            print(f"\n  ⚠️  批次 {i // batch_size + 1} 失败: {e}")
            time.sleep(1)
            continue
        done = min(i + batch_size, len(pmids))
        print(f"  进度：{done}/{len(pmids)} ({100*done//max(len(pmids),1)}%)  已解析：{len(records)} 篇",
              end="\r", flush=True)
        time.sleep(sleep_sec)
    print(f"\n  ✅ 解析 {len(records)} 篇文章")

    return records


# ---------------------------------------------------------------------------
# 导出
# ---------------------------------------------------------------------------

def export_excel(records: List[Dict], out_file: str):
    """导出格式化 Excel，与 pubmed-improved.py 样式一致，增加新列。"""
    df = pd.DataFrame(records)

    # 按日期排序
    df["_sort"] = pd.to_datetime(df["Publication Date"], errors="coerce")
    df = df.sort_values("_sort", ascending=False).drop(columns=["_sort"])
    df.reset_index(drop=True, inplace=True)
    df.index += 1

    # 列顺序
    col_order = [
        "PMID", "Title", "Abstract", "First Author", "All Authors",
        "Journal", "Journal Abbr", "Volume", "Issue", "Pages",
        "Publication Date", "DOI", "PMCID",
        "Keywords", "MeSH Terms", "Publication Types", "Language",
        "PubMed URL", "PMC URL", "DOI URL", "Full Text URL",
    ]
    df = df[[c for c in col_order if c in df.columns]]

    os.makedirs(os.path.dirname(out_file) if os.path.dirname(out_file) else ".", exist_ok=True)

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=True, index_label="No.", sheet_name="Results")
        ws = writer.sheets["Results"]

        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

        # 列宽配置
        col_widths = {
            "A": 6,  "B": 12, "C": 60, "D": 80, "E": 20,
            "F": 40, "G": 25, "H": 20, "I": 8,  "J": 8,
            "K": 8,  "L": 14, "M": 12, "N": 12,
            "O": 40, "P": 50, "Q": 20, "R": 8,
            "S": 40, "T": 40, "U": 40, "V": 40,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        even_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        odd_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.fill = even_fill if row_idx % 2 == 0 else odd_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row_idx].height = 50

    print(f"  📊 Excel: {out_file}（{len(df)} 篇）")


def export_json(records: List[Dict], out_file: str, meta: Dict):
    """
    导出 JSON 文件。
    结构：
    {
      "meta": { 检索参数、时间、数量 },
      "articles": [ { pmid, pmcid, doi, title, abstract, authors, journal,
                       publication_date, keywords, mesh_terms, links, ... } ]
    }
    """
    entries = [record_to_json_entry(r) for r in records]

    output = {
        "meta": {
            **meta,
            "total": len(entries),
        },
        "articles": entries,
    }

    os.makedirs(os.path.dirname(out_file) if os.path.dirname(out_file) else ".", exist_ok=True)
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"  📄 JSON:  {out_file}（{len(entries)} 条）")


# ---------------------------------------------------------------------------
# 主函数
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="PubMed 文献检索工具（增强版）")
    parser.add_argument("--config", default="config.json", help="配置文件路径")
    args = parser.parse_args()

    print("=" * 70)
    print("  PubMed 文献检索工具（增强版）")
    print("=" * 70)

    cfg = load_config(args.config)

    query      = cfg["query"]
    date_start = cfg.get("date_start", "2017")
    date_end   = cfg.get("date_end",   "2025")
    api_key    = cfg.get("api_key")
    batch_size = int(cfg.get("batch_size", 100))
    sleep_sec  = float(cfg.get("sleep_sec", 0.4))

    # 输出路径：xlsx 来自 config，json 同目录同文件名
    out_xlsx = cfg.get("out_file", "output/pubmed_results.xlsx")
    base     = os.path.splitext(out_xlsx)[0]
    out_json = cfg.get("out_json", base + ".json")

    print(f"  时间范围  : {date_start} ~ {date_end}")
    print(f"  检索词    : {query[:80]}...")
    print(f"  输出 xlsx : {out_xlsx}")
    print(f"  输出 json : {out_json}")

    records = fetch_records(query, date_start, date_end, api_key, batch_size, sleep_sec)

    if not records:
        print("\n⚠️  未检索到任何文献")
        return

    import datetime
    meta = {
        "query":      query,
        "date_start": date_start,
        "date_end":   date_end,
        "fetched_at": datetime.datetime.now().isoformat(timespec="seconds"),
    }

    print("\n导出结果...")
    export_excel(records, out_xlsx)
    export_json(records, out_json, meta)

    print(f"\n{'=' * 70}")
    print(f"  ✅ 完成！共 {len(records)} 篇文献")
    print(f"{'=' * 70}\n")


if __name__ == "__main__":
    main()
