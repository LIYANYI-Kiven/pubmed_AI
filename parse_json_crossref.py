#!/usr/bin/env python3
"""
parse_json_crossref.py

整理 crossref_search.py 生成的 JSON 文件（原始格式或解析格式均可）。
提取字段：标题、PubMed ID（Crossref 中可能缺失）、DOI、PubMed 摘要链接。

Crossref 本身不包含 PMID，如果需要 PMID 本脚本会尝试通过
NCBI ID Converter API（doi → pmid）批量补全。

用法:
    python parse_json_crossref.py
    python parse_json_crossref.py --input output/crossref_raw.json --output output/crossref_summary.xlsx
    python parse_json_crossref.py --no-pmid-lookup   # 跳过 PMID 查询，速度更快
"""

import argparse
import json
import os
import time
from typing import Dict, List, Optional

import requests
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


# ---------------------------------------------------------------------------
# NCBI ID Converter：DOI → PMID
# ---------------------------------------------------------------------------

def doi_to_pmid_batch(dois: List[str], api_key: Optional[str] = None,
                       batch_size: int = 50, sleep_sec: float = 0.5) -> Dict[str, str]:
    """
    通过 NCBI ID Converter API 将 DOI 列表转换为 PMID。
    返回 {doi: pmid}，找不到则无对应键。
    """
    if not dois:
        return {}

    url = "https://www.ncbi.nlm.nih.gov/pmc/utils/idconv/v1.0/"
    params_base = {}
    if api_key:
        params_base["api_key"] = api_key

    result = {}
    session = requests.Session()
    session.headers["User-Agent"] = "parse-crossref/1.0"

    for i in range(0, len(dois), batch_size):
        chunk = [str(d) for d in dois[i:i + batch_size] if d]  # 确保全为字符串
        try:
            r = session.get(
                url,
                params={
                    "ids":    ",".join(chunk),
                    "format": "json",
                    "tool":   "parse_crossref",
                    **params_base,
                },
                timeout=30,
            )
            r.raise_for_status()
            data = r.json()
            for record in data.get("records", []):
                doi  = (record.get("doi") or "").strip()
                pmid = str(record.get("pmid") or "").strip()
                if doi and pmid:
                    result[doi.lower()] = pmid
        except Exception as e:
            print(f"  ⚠️  PMID 查询失败 (batch {i}): {e}")

        time.sleep(sleep_sec)

    return result


# ---------------------------------------------------------------------------
# 解析 Crossref JSON
# ---------------------------------------------------------------------------

def _get_pub_date(item: Dict) -> str:
    """从 Crossref 条目中提取发表日期。"""
    for key in ["published", "published-print", "published-online", "created"]:
        dp = item.get(key, {}).get("date-parts", [[]])
        if dp and dp[0]:
            parts = dp[0]
            if len(parts) >= 3:
                return f"{parts[0]}-{parts[1]:02d}-{parts[2]:02d}"
            elif len(parts) == 2:
                return f"{parts[0]}-{parts[1]:02d}"
            elif len(parts) == 1:
                return str(parts[0])
    return ""


def parse_crossref_json(input_file: str, lookup_pmid: bool = True,
                        api_key: Optional[str] = None) -> List[Dict]:
    """
    从 crossref_search.py 生成的 JSON 中提取核心字段。
    支持两种格式：
      - 原始 API 格式：列表，每条是完整 Crossref work 对象
      - 解析后格式：列表，每条已含 DOI/Title/Published 等扁平字段
    """
    with open(input_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        print("⚠️  JSON 格式不是列表，尝试取 message.items")
        data = data.get("message", {}).get("items", [])

    if not data:
        return []

    # 判断是原始格式还是解析后格式
    first = data[0]
    is_raw = "DOI" in first and isinstance(first.get("title", ""), list)
    is_parsed = "DOI" in first and isinstance(first.get("Title", ""), str)

    items_parsed = []
    for item in data:
        if is_raw:
            # 原始 Crossref API 格式
            titles = item.get("title", [])
            title  = titles[0] if titles else ""
            doi    = (item.get("DOI") or "").strip()
            pub_date = _get_pub_date(item)
            # PMID 从 relation 或 link 里尝试提取
            pmid = ""
            for link in item.get("link", []):
                href = link.get("URL", "")
                if "pubmed" in href.lower() or "nlm.nih.gov" in href.lower():
                    pmid = href.rstrip("/").split("/")[-1]
                    if pmid.isdigit():
                        break
                    pmid = ""
        elif is_parsed:
            # crossref_search.py parse_item() 输出的扁平格式
            title    = item.get("Title", "")
            doi      = (item.get("DOI") or "").strip()
            pub_date = item.get("Published", "")
            pmid     = (item.get("PMID") or "").strip()
        else:
            # 未知格式，尽量提取
            title    = str(item.get("title") or item.get("Title") or "")
            doi      = str(item.get("DOI") or item.get("doi") or "").strip()
            pub_date = ""
            pmid     = ""

        items_parsed.append({
            "title":    title,
            "doi":      doi,
            "pmid":     pmid,
            "pub_date": pub_date,
        })

    # 批量通过 DOI 查询缺失的 PMID
    if lookup_pmid:
        missing_dois = [x["doi"] for x in items_parsed if x["doi"] and not x["pmid"]]
        if missing_dois:
            print(f"  通过 NCBI ID Converter 查询 {len(missing_dois)} 个 DOI 对应的 PMID...")
            doi_pmid_map = doi_to_pmid_batch(missing_dois, api_key=api_key)
            for item in items_parsed:
                if not item["pmid"] and item["doi"]:
                    item["pmid"] = doi_pmid_map.get(item["doi"].lower(), "")
        else:
            print("  所有条目已有 PMID，跳过查询")

    # 构建最终记录
    records = []
    for i, item in enumerate(items_parsed, 1):
        pmid       = item["pmid"]
        doi        = item["doi"]
        pubmed_url = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else ""
        doi_url    = f"https://doi.org/{doi}" if doi else ""

        records.append({
            "No.":             i,
            "Title":           item["title"],
            "PubMed ID":       pmid,
            "DOI":             doi,
            "PubMed Abstract": pubmed_url,
            "DOI URL":         doi_url,
            "Publication Date": item["pub_date"],
        })

    return records


# ---------------------------------------------------------------------------
# 导出
# ---------------------------------------------------------------------------

def export_excel(records: List[Dict], out_file: str):
    df = pd.DataFrame(records).set_index("No.")
    os.makedirs(os.path.dirname(out_file) if os.path.dirname(out_file) else ".", exist_ok=True)

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Summary")
        ws = writer.sheets["Summary"]

        fill = PatternFill(start_color="1B6CA8", end_color="1B6CA8", fill_type="solid")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        col_widths = {"A": 6, "B": 70, "C": 12, "D": 35, "E": 45, "F": 45, "G": 14}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        even = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.fill = even if row_idx % 2 == 0 else odd
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row_idx].height = 40

    print(f"  📊 Excel: {out_file}（{len(df)} 篇）")


# ---------------------------------------------------------------------------
# 主函数
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="整理 Crossref JSON，补全 PMID")
    parser.add_argument("--input",        default="output/crossref_raw.json",     help="输入 JSON 路径")
    parser.add_argument("--output",       default="output/crossref_summary.xlsx", help="输出 Excel 路径")
    parser.add_argument("--no-pmid-lookup", action="store_true", help="跳过 DOI→PMID 查询")
    parser.add_argument("--api-key",      default=None, help="NCBI API Key（可选）")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"❌ 文件不存在: {args.input}")
        return

    print(f"读取: {args.input}")
    records = parse_crossref_json(
        args.input,
        lookup_pmid=not args.no_pmid_lookup,
        api_key=args.api_key
    )

    if not records:
        print("⚠️  无有效记录")
        return

    print(f"共 {len(records)} 篇文献")
    export_excel(records, args.output)

    out_json = os.path.splitext(args.output)[0] + "_detail.json"
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    print(f"  📄 JSON:  {out_json}")


if __name__ == "__main__":
    main()
