#!/usr/bin/env python3
"""
PubMed 全文版（V2）

在版本 1（检索式 + 摘要筛选）基础上升级，版本 1 脚本保持不变。

V2 流程：
1. 按 config_v2.json 中的检索式从 PubMed 获取文献列表（也可额外指定 seed_pmids）；
2. 对每篇文献依次尝试获取全文：
   a. 下载开放获取 PDF → 提取文本 → 提交 AI（文献判断 + 参考文献挖掘）；
   b. 若 PDF 不可用 → 访问 PMC 在线全文（XML / 网页 HTML）→ 同样提交 AI；
   c. 若仍无法获取全文 → 回退到版本 1 逻辑，仅基于摘要判断。
3. 导出 Excel：文献筛选结果 + 参考文献挖掘结果。

配置见 config_v2.json。AI key 优先读环境变量 AI_API_KEY。
"""

import json
import os
import time
import re
from typing import Dict, List, Optional, Tuple
import xml.etree.ElementTree as ET

import requests
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from pypdf import PdfReader
except ImportError:
    PdfReader = None


# ---------------------------------------------------------------------------
# 配置
# ---------------------------------------------------------------------------

PROMPT_FILE_MAP = [
    ("system_prompt", "system_prompt_file"),
    ("user_prompt_template", "user_prompt_file"),
    ("pdf_system_prompt", "pdf_system_prompt_file"),
    ("pdf_user_prompt_template", "pdf_user_prompt_file"),
    ("abstract_system_prompt", "abstract_system_prompt_file"),
    ("abstract_user_prompt_template", "abstract_user_prompt_file"),
    ("article_fulltext_system_prompt", "article_fulltext_system_prompt_file"),
    ("article_fulltext_user_prompt_template", "article_fulltext_user_prompt_file"),
    ("pdf_refs_system_prompt", "pdf_refs_system_prompt_file"),
    ("pdf_refs_user_prompt_template", "pdf_refs_user_prompt_file"),
]


def load_config(path: str = "config_v2.json") -> Dict:
    if not os.path.exists(path):
        raise FileNotFoundError(f"配置文件不存在: {path}")
    with open(path, "r", encoding="utf-8") as f:
        config = json.load(f)

    for key, file_key in PROMPT_FILE_MAP:
        fp = config.get(file_key)
        if fp:
            if not os.path.exists(fp):
                raise FileNotFoundError(f"提示词文件不存在: {fp}（配置项: {file_key}）")
            with open(fp, "r", encoding="utf-8") as f:
                config[key] = f.read()
    return config


# ---------------------------------------------------------------------------
# NCBI E-utils 客户端
# ---------------------------------------------------------------------------

class NCBIClient:
    BASE = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"

    def __init__(self, api_key: Optional[str] = None, sleep_sec: float = 0.4):
        self.params = {"api_key": api_key} if api_key else {}
        self.sleep_sec = sleep_sec
        self.session = requests.Session()
        self.session.headers.update({"User-Agent": "pubmed-fulltext-v2/2.0"})

    def search_with_history(self, query: str, date_start: str, date_end: str,
                            timeout: float = 30.0) -> Dict:
        r = self.session.get(
            self.BASE + "esearch.fcgi",
            params={
                "db": "pubmed", "term": query,
                "datetype": "pdat", "mindate": date_start, "maxdate": date_end,
                "retmode": "json", "retmax": 0, "usehistory": "y",
                "sort": "pub_date", **self.params,
            },
            timeout=timeout,
        )
        r.raise_for_status()
        result = r.json()["esearchresult"]
        return {
            "count": int(result["count"]),
            "webenv": result.get("webenv"),
            "query_key": result.get("querykey"),
        }

    def get_pmids_with_history(self, webenv: str, query_key: str,
                               batch_size: int, start: int,
                               timeout: float = 30.0) -> List[str]:
        r = self.session.get(
            self.BASE + "esearch.fcgi",
            params={
                "db": "pubmed", "retmode": "json",
                "retmax": min(batch_size, 10000), "retstart": start,
                "webenv": webenv, "query_key": query_key, **self.params,
            },
            timeout=timeout,
        )
        r.raise_for_status()
        return r.json()["esearchresult"].get("idlist", [])

    def pmid_to_pmcid(self, pmid: str, timeout: float = 30.0) -> Optional[str]:
        r = self.session.get(
            self.BASE + "elink.fcgi",
            params={"dbfrom": "pubmed", "db": "pmc", "id": pmid,
                    "retmode": "json", **self.params},
            timeout=timeout,
        )
        r.raise_for_status()
        try:
            linksets = r.json()["linksets"][0].get("linksetdbs", [])
            for ls in linksets:
                if ls.get("dbto") == "pmc" and ls.get("links"):
                    return "PMC" + str(ls["links"][0])
        except (KeyError, IndexError):
            return None
        return None

    def fetch_pmc_xml(self, pmcid: str, timeout: float = 60.0) -> ET.Element:
        pmc_num = pmcid.replace("PMC", "")
        r = self.session.get(
            self.BASE + "efetch.fcgi",
            params={"db": "pmc", "id": pmc_num, "retmode": "xml", **self.params},
            timeout=timeout,
        )
        r.raise_for_status()
        return ET.fromstring(r.content)

    def fetch_pubmed_summaries(self, pmids: List[str],
                               timeout: float = 60.0) -> Dict[str, Dict]:
        if not pmids:
            return {}
        r = self.session.get(
            self.BASE + "efetch.fcgi",
            params={"db": "pubmed", "id": ",".join(pmids),
                    "retmode": "xml", "rettype": "abstract", **self.params},
            timeout=timeout,
        )
        r.raise_for_status()
        root = ET.fromstring(r.content)
        out: Dict[str, Dict] = {}
        for art in root.findall(".//PubmedArticle"):
            out[art.findtext(".//PMID", "").strip()] = parse_pubmed_article(art)
        return out

    def fetch_pubmed_article(self, pmid: str, timeout: float = 60.0) -> Dict[str, str]:
        summaries = self.fetch_pubmed_summaries([pmid], timeout=timeout)
        return summaries.get(pmid, {"PMID": pmid, "Title": "", "Abstract": "", "DOI": ""})


def parse_pubmed_article(article: ET.Element) -> Dict[str, str]:
    pmid = article.findtext(".//PMID", "").strip()
    title_elem = article.find(".//ArticleTitle")
    title = "".join(title_elem.itertext()) if title_elem is not None else ""
    abstract = " ".join(
        (f"[{a.attrib.get('Label', '')}] " if a.attrib.get("Label") else "")
        + (a.text or "")
        for a in article.findall(".//AbstractText")
        if a.text
    ).strip()
    doi = ""
    for aid in article.findall(".//ArticleId"):
        if aid.attrib.get("IdType") == "doi":
            doi = (aid.text or "").strip()
            break
    return {"PMID": pmid, "Title": title, "Abstract": abstract, "DOI": doi}


def search_pmids(cfg: Dict, ncbi: NCBIClient) -> List[str]:
    query = cfg.get("query", "").strip()
    if not query:
        return []

    date_start = cfg.get("date_start", "2020")
    date_end = cfg.get("date_end", "2025")
    batch_size = int(cfg.get("search_batch_size", 100))
    max_articles = cfg.get("max_articles")

    print(f"\n[检索] PubMed 检索式: {query[:80]}...")
    print(f"  时间范围: {date_start} ~ {date_end}")

    search = ncbi.search_with_history(query, date_start, date_end)
    total = search["count"]
    print(f"  命中: {total} 篇")

    if total == 0:
        return []

    limit = min(total, max_articles) if max_articles else total
    pmids: List[str] = []
    for start in range(0, limit, batch_size):
        batch = ncbi.get_pmids_with_history(
            search["webenv"], search["query_key"],
            min(batch_size, limit - start), start,
        )
        pmids.extend(batch)
        print(f"  进度: {len(pmids)}/{limit}", end="\r", flush=True)
        time.sleep(ncbi.sleep_sec)
    print(f"\n  完成: 获取 {len(pmids)} 个 PMID")
    return pmids


# ---------------------------------------------------------------------------
# PDF 链接查找 / 下载 / 文本提取
# ---------------------------------------------------------------------------

class PDFFetcher:
    PMC_OA = "https://www.ncbi.nlm.nih.gov/pmc/utils/oa/oa.fcgi"
    UNPAYWALL = "https://api.unpaywall.org/v2/"

    def __init__(self, ncbi: NCBIClient, pdf_dir: str,
                 unpaywall_email: Optional[str] = None,
                 timeout: float = 60.0):
        self.ncbi = ncbi
        self.pdf_dir = pdf_dir
        self.unpaywall_email = unpaywall_email
        self.timeout = timeout
        self.session = ncbi.session

    def find_pdf_url(self, pmid: str, doi: str) -> Optional[str]:
        url = self._pmc_oa_pdf(pmid) if pmid else None
        if not url and doi and self.unpaywall_email:
            url = self._unpaywall_pdf(doi)
        return url

    def _pmc_oa_pdf(self, pmid: str) -> Optional[str]:
        pmcid = self.ncbi.pmid_to_pmcid(pmid)
        time.sleep(self.ncbi.sleep_sec)
        if not pmcid:
            return None
        try:
            r = self.session.get(self.PMC_OA, params={"id": pmcid}, timeout=self.timeout)
            r.raise_for_status()
            root = ET.fromstring(r.content)
            link = root.find(".//record/link[@format='pdf']")
            if link is not None:
                return link.attrib.get("href", "").replace("ftp://", "https://")
        except (requests.RequestException, ET.ParseError):
            return None
        return None

    def _unpaywall_pdf(self, doi: str) -> Optional[str]:
        try:
            r = self.session.get(self.UNPAYWALL + doi,
                                 params={"email": self.unpaywall_email},
                                 timeout=self.timeout)
            r.raise_for_status()
            loc = r.json().get("best_oa_location") or {}
            return loc.get("url_for_pdf")
        except (requests.RequestException, ValueError):
            return None

    def download(self, url: str, pmid: str, subdir: str = "source") -> Optional[str]:
        dest_dir = os.path.join(self.pdf_dir, subdir)
        os.makedirs(dest_dir, exist_ok=True)
        path = os.path.join(dest_dir, f"{pmid}.pdf")
        if os.path.exists(path) and os.path.getsize(path) > 0:
            return path
        try:
            r = self.session.get(url, timeout=self.timeout, stream=True)
            r.raise_for_status()
            ctype = r.headers.get("Content-Type", "")
            first = next(r.iter_content(chunk_size=1024), b"")
            if "pdf" not in ctype.lower() and not first.startswith(b"%PDF"):
                return None
            with open(path, "wb") as f:
                f.write(first)
                for chunk in r.iter_content(chunk_size=8192):
                    f.write(chunk)
            return path
        except requests.RequestException:
            return None

    @staticmethod
    def extract_text(path: str, max_len: int = 30000) -> str:
        if PdfReader is None:
            raise RuntimeError("缺少 pypdf，请先运行: pip install pypdf")
        try:
            reader = PdfReader(path)
        except Exception:
            return ""
        parts, total = [], 0
        for page in reader.pages:
            try:
                txt = page.extract_text() or ""
            except Exception:
                txt = ""
            parts.append(txt)
            total += len(txt)
            if total >= max_len:
                break
        text = re.sub(r"\s+\n", "\n", "\n".join(parts)).strip()
        return text[:max_len]


# ---------------------------------------------------------------------------
# 全文获取：PDF → PMC 网页 → 摘要
# ---------------------------------------------------------------------------

def _itertext(elem: Optional[ET.Element]) -> str:
    return re.sub(r"\s+", " ", "".join(elem.itertext())).strip() if elem is not None else ""


def _find_or_root(root: ET.Element, path: str) -> ET.Element:
    elem = root.find(path)
    return elem if elem is not None else root


def parse_fulltext(root: ET.Element) -> Dict[str, str]:
    article = _find_or_root(root, ".//article")
    title = _itertext(article.find(".//article-meta//article-title"))
    abstract = _itertext(article.find(".//article-meta//abstract"))
    body = article.find(".//body")
    if body is not None:
        paras = [_itertext(p) for p in body.findall(".//p")]
        fulltext = "\n".join(p for p in paras if p)
    else:
        fulltext = ""
    return {"Title": title, "Abstract": abstract, "FullText": fulltext}


def _find_citation(ref: ET.Element) -> ET.Element:
    for path in (".//element-citation", ".//mixed-citation", ".//citation"):
        elem = ref.find(path)
        if elem is not None:
            return elem
    return ref


def parse_references(root: ET.Element) -> List[Dict[str, str]]:
    article = _find_or_root(root, ".//article")
    refs: List[Dict[str, str]] = []
    for idx, ref in enumerate(article.findall(".//ref-list//ref"), 1):
        citation = _find_citation(ref)
        title = _itertext(citation.find(".//article-title"))
        source = _itertext(citation.find(".//source"))
        year = _itertext(citation.find(".//year"))
        authors = []
        for name in citation.findall(".//name"):
            surname = _itertext(name.find("surname"))
            given = _itertext(name.find("given-names"))
            if surname:
                authors.append(f"{surname} {given}".strip())
        author_str = "; ".join(authors[:3]) + (" et al." if len(authors) > 3 else "")
        pmid, doi = "", ""
        for pid in citation.findall(".//pub-id"):
            ptype = pid.attrib.get("pub-id-type", "")
            if ptype == "pmid":
                pmid = (pid.text or "").strip()
            elif ptype == "doi":
                doi = (pid.text or "").strip()
        if not title:
            title = _itertext(citation)
        refs.append({
            "Ref Index": idx, "PMID": pmid, "DOI": doi,
            "Title": title, "Authors": author_str,
            "Source": source, "Year": year, "Abstract": "",
        })
    return refs


def fetch_pmc_html_text(ncbi: NCBIClient, pmcid: str, timeout: float = 60.0) -> str:
    """抓取 PMC 网页正文文本（PDF/XML 均不可用时的备选）。"""
    url = f"https://www.ncbi.nlm.nih.gov/pmc/articles/{pmcid}/"
    try:
        r = ncbi.session.get(url, timeout=timeout)
        r.raise_for_status()
        html = r.text
        html = re.sub(r"<script[^>]*>.*?</script>", " ", html, flags=re.DOTALL | re.I)
        html = re.sub(r"<style[^>]*>.*?</style>", " ", html, flags=re.DOTALL | re.I)
        text = re.sub(r"<[^>]+>", " ", html)
        text = re.sub(r"\s+", " ", text).strip()
        return text
    except requests.RequestException:
        return ""


def enrich_references(refs: List[Dict[str, str]], client: NCBIClient,
                      batch_size: int = 100, verbose: bool = True) -> None:
    pmids = [r["PMID"] for r in refs if r["PMID"]]
    if not pmids:
        if verbose:
            print("  ℹ️  参考文献中未发现 PMID，跳过摘要补全")
        return
    summaries: Dict[str, Dict] = {}
    for i in range(0, len(pmids), batch_size):
        chunk = pmids[i:i + batch_size]
        try:
            summaries.update(client.fetch_pubmed_summaries(chunk))
        except (requests.RequestException, ET.ParseError) as e:
            if verbose:
                print(f"  ⚠️  补全批次失败: {e}")
        time.sleep(client.sleep_sec)
    for r in refs:
        info = summaries.get(r["PMID"])
        if info:
            r["Abstract"] = info.get("Abstract", "")
            if not r["Title"]:
                r["Title"] = info.get("Title", "")
    if verbose:
        print(f"  ✅ 已为 {len(summaries)}/{len(pmids)} 篇含 PMID 的参考文献补全题录")


class AcquiredContent:
    """单篇文献获取到的内容与元数据。"""

    def __init__(self):
        self.source = "abstract"          # pdf | pmc_web | pmc_html | abstract
        self.fulltext = ""
        self.pmcid = ""
        self.pdf_url = ""
        self.pdf_file = ""
        self.refs: List[Dict[str, str]] = []
        self.meta: Dict[str, str] = {}


def acquire_content(pmid: str, ncbi: NCBIClient,
                    fetcher: Optional[PDFFetcher],
                    cfg: Dict) -> AcquiredContent:
    """
    按优先级获取全文：PDF → PMC XML → PMC 网页 → 仅摘要。
    """
    ac = AcquiredContent()
    ac.meta = ncbi.fetch_pubmed_article(pmid)
    time.sleep(ncbi.sleep_sec)

    doi = ac.meta.get("DOI", "")
    pdf_text_max = cfg.get("pdf_text_max_len", 30000)
    fulltext_max = cfg.get("fulltext_max_len", 12000)

    # --- 1. 尝试 PDF ---
    if fetcher is not None and cfg.get("download_pdf", True):
        print("  [1/3] 尝试下载 PDF...", end=" ", flush=True)
        pdf_url = fetcher.find_pdf_url(pmid, doi)
        if pdf_url:
            path = fetcher.download(pdf_url, pmid, subdir="source")
            if path:
                text = fetcher.extract_text(path, pdf_text_max)
                if len(text) >= 200:
                    ac.source = "pdf"
                    ac.fulltext = text[:fulltext_max]
                    ac.pdf_url = pdf_url
                    ac.pdf_file = path
                    print(f"成功 ({len(text)} 字符)")
                    # 同时尝试 PMC 获取结构化参考文献
                    pmcid = ncbi.pmid_to_pmcid(pmid)
                    if pmcid:
                        ac.pmcid = pmcid
                        try:
                            root = ncbi.fetch_pmc_xml(pmcid)
                            ac.refs = parse_references(root)
                            time.sleep(ncbi.sleep_sec)
                        except (requests.RequestException, ET.ParseError):
                            pass
                    return ac
                print("文本过短")
            else:
                print("下载失败", end=" ")
        else:
            print("无链接", end=" ")
        print("→ 尝试网页全文")

    # --- 2. 尝试 PMC 在线全文 ---
    print("  [2/3] 尝试 PMC 在线全文...", end=" ", flush=True)
    pmcid = ncbi.pmid_to_pmcid(pmid)
    time.sleep(ncbi.sleep_sec)
    if pmcid:
        ac.pmcid = pmcid
        try:
            root = ncbi.fetch_pmc_xml(pmcid)
            ft = parse_fulltext(root)
            ac.refs = parse_references(root)
            body = ft["FullText"]
            if len(body) >= 200:
                ac.source = "pmc_web"
                title = ft["Title"] or ac.meta.get("Title", "")
                abstract = ft["Abstract"] or ac.meta.get("Abstract", "")
                combined = f"{title}\n\n{abstract}\n\n{body}".strip()
                ac.fulltext = combined[:fulltext_max]
                if ac.meta.get("Title") == "":
                    ac.meta["Title"] = title
                print(f"XML 成功 ({len(body)} 字符, {len(ac.refs)} 条参考文献)")
                return ac
        except (requests.RequestException, ET.ParseError):
            pass

        html_text = fetch_pmc_html_text(ncbi, pmcid)
        time.sleep(ncbi.sleep_sec)
        if len(html_text) >= 500:
            ac.source = "pmc_html"
            ac.fulltext = html_text[:fulltext_max]
            print(f"网页成功 ({len(html_text)} 字符)")
            return ac
        print("网页文本过短", end=" ")
    else:
        print("无 PMC 记录", end=" ")
    print("→ 回退摘要")

    # --- 3. 仅摘要（版本 1 逻辑）---
    print("  [3/3] 使用 PubMed 摘要")
    ac.source = "abstract"
    ac.fulltext = ""
    return ac


# ---------------------------------------------------------------------------
# AI 分析
# ---------------------------------------------------------------------------

class ArticleAnalyzer:
    """文献判断 + 参考文献筛选。"""

    def __init__(self, cfg: Dict, api_key: str):
        self.api_url = cfg.get("ai_api_url")
        self.model = cfg.get("ai_model", "deepseek-chat")
        self.result_key = cfg.get("result_key", "is_relevant")
        self.fulltext_max_len = cfg.get("fulltext_max_len", 12000)
        self.ref_abstract_max_len = cfg.get("ref_abstract_max_len", 500)
        self.pdf_text_max_len = cfg.get("pdf_text_max_len", 30000)
        self.headers = {"Content-Type": "application/json",
                        "Authorization": f"Bearer {api_key}"}

        self.system_prompt = cfg.get("system_prompt", "")
        self.user_prompt_template = cfg.get("user_prompt_template", "")
        self.pdf_system_prompt = cfg.get("pdf_system_prompt", "")
        self.pdf_user_prompt_template = cfg.get("pdf_user_prompt_template", "")
        self.abstract_system_prompt = cfg.get("abstract_system_prompt", "")
        self.abstract_user_prompt_template = cfg.get("abstract_user_prompt_template", "")
        self.article_fulltext_system_prompt = cfg.get("article_fulltext_system_prompt", "")
        self.article_fulltext_user_prompt_template = cfg.get(
            "article_fulltext_user_prompt_template", "")
        self.pdf_refs_system_prompt = cfg.get("pdf_refs_system_prompt", "")
        self.pdf_refs_user_prompt_template = cfg.get("pdf_refs_user_prompt_template", "")

    def _chat(self, system_prompt: str, user_prompt: str, timeout: float = 180) -> str:
        resp = requests.post(
            self.api_url, headers=self.headers,
            json={"model": self.model,
                  "messages": [{"role": "system", "content": system_prompt},
                               {"role": "user", "content": user_prompt}],
                  "temperature": 0.3},
            timeout=timeout,
        )
        resp.raise_for_status()
        content = resp.json()["choices"][0]["message"]["content"].strip()
        if content.startswith("```json"):
            content = content[7:]
        elif content.startswith("```"):
            content = content[3:]
        if content.endswith("```"):
            content = content[:-3]
        return content.strip()

    def judge_article_fulltext(self, pmid: str, title: str, fulltext: str) -> Dict:
        text = fulltext[:self.fulltext_max_len]
        prompt = (self.article_fulltext_user_prompt_template
                  .replace("{pmid}", pmid or "N/A")
                  .replace("{title}", title or "N/A")
                  .replace("{fulltext}", text))
        try:
            content = self._chat(self.article_fulltext_system_prompt, prompt)
            return json.loads(content)
        except Exception as e:
            return {"is_relevant": None, "confidence": "",
                    "reason": f"全文判断失败: {type(e).__name__}: {e}"}

    def judge_article_abstract(self, pmid: str, title: str, abstract: str) -> Dict:
        """版本 1 逻辑：仅基于摘要判断。"""
        abs_text = abstract or "N/A"
        if len(abs_text) > self.ref_abstract_max_len:
            abs_text = abs_text[:self.ref_abstract_max_len] + "..."
        prompt = (self.abstract_user_prompt_template
                  .replace("{pmid}", pmid or "N/A")
                  .replace("{title}", title or "N/A")
                  .replace("{abstract}", abs_text))
        try:
            content = self._chat(self.abstract_system_prompt, prompt)
            return json.loads(content)
        except Exception as e:
            return {"is_relevant": None, "confidence": "",
                    "reason": f"摘要判断失败: {type(e).__name__}: {e}"}

    def find_refs_in_pdf_text(self, pmid: str, title: str, fulltext: str) -> List[Dict]:
        """从 PDF 全文中由 AI 挖掘可能相关的参考文献（无结构化 ref 列表时）。"""
        text = fulltext[:self.pdf_text_max_len]
        prompt = (self.pdf_refs_user_prompt_template
                  .replace("{pmid}", pmid or "N/A")
                  .replace("{title}", title or "N/A")
                  .replace("{fulltext}", text))
        try:
            content = self._chat(self.pdf_refs_system_prompt, prompt)
            data = json.loads(content)
            refs = data.get("references", [])
            for r in refs:
                r.setdefault("Source Type", "pdf_text_mining")
            return refs
        except Exception as e:
            print(f"  ⚠️  PDF 参考文献挖掘失败: {e}")
            return []

    def _build_ref_prompt(self, fulltext: str, refs: List[Dict]) -> str:
        ft = fulltext[:self.fulltext_max_len]
        if len(fulltext) > self.fulltext_max_len:
            ft += "..."
        lines = []
        for r in refs:
            abs = r.get("Abstract", "") or "N/A"
            if isinstance(abs, str) and len(abs) > self.ref_abstract_max_len:
                abs = abs[:self.ref_abstract_max_len] + "..."
            lines.append(
                f"[{r.get('Ref Index')}] PMID:{r.get('PMID') or 'N/A'}\n"
                f"Title: {r.get('Title', 'N/A')}\n"
                f"Source: {r.get('Source', '')} ({r.get('Year', '')})\n"
                f"Abstract: {abs}\n"
            )
        prompt = self.user_prompt_template
        prompt = prompt.replace("{fulltext}", ft)
        prompt = prompt.replace("{n}", str(len(refs)))
        prompt = prompt.replace("{articles}", "".join(lines))
        return prompt

    def screen_ref_batch(self, fulltext: str, refs: List[Dict], batch_num: int) -> Dict:
        prompt = self._build_ref_prompt(fulltext, refs)
        try:
            content = self._chat(self.system_prompt, prompt)
            results = json.loads(content).get("articles", [])
            return {"batch_num": batch_num, "success": True, "results": results}
        except Exception as e:
            print(f"  ⚠️  参考文献批次 {batch_num} 失败: {e}")
            return {"batch_num": batch_num, "success": False, "results": []}

    def review_ref_pdf(self, pmid: str, title: str, pdf_text: str) -> Dict:
        text = pdf_text[:self.pdf_text_max_len]
        prompt = (self.pdf_user_prompt_template
                  .replace("{pmid}", pmid or "N/A")
                  .replace("{title}", title or "N/A")
                  .replace("{pdftext}", text))
        try:
            content = self._chat(self.pdf_system_prompt, prompt)
            return json.loads(content)
        except Exception as e:
            return {"is_relevant": None, "confidence": "",
                    "reason": f"PDF 复查失败: {type(e).__name__}: {e}"}


def screen_structured_refs(analyzer: ArticleAnalyzer, fulltext: str,
                           refs: List[Dict], batch_size: int, sleep_sec: float,
                           result_key: str) -> List[Dict]:
    by_index = {r["Ref Index"]: r for r in refs}
    by_pmid = {r["PMID"]: r for r in refs if r["PMID"]}
    kept: List[Dict] = []
    batches = [refs[i:i + batch_size] for i in range(0, len(refs), batch_size)]
    for i, batch in enumerate(batches, 1):
        print(f"    参考文献批次 {i}/{len(batches)}...", end=" ", flush=True)
        res = analyzer.screen_ref_batch(fulltext, batch, i)
        print("OK" if res["success"] else "FAIL")
        for item in res["results"]:
            if not item.get(result_key):
                continue
            ref = by_index.get(item.get("ref_index")) or by_pmid.get(str(item.get("pmid", "")))
            if ref is None:
                continue
            rec = dict(ref)
            rec["AI_Confidence"] = item.get("confidence", "")
            rec["AI_Reason"] = item.get("reason", "")
            rec["Source Type"] = "structured"
            kept.append(rec)
        if i < len(batches):
            time.sleep(sleep_sec)
    return kept


def pdf_review_refs(kept: List[Dict], fetcher: PDFFetcher,
                    analyzer: ArticleAnalyzer, sleep_sec: float) -> None:
    for r in kept:
        r.setdefault("PDF_URL", "")
        r.setdefault("PDF_File", "")
        r.setdefault("PDF_Relevant", "")
        r.setdefault("PDF_Confidence", "")
        r.setdefault("PDF_Reason", "")
        pmid, doi = r.get("PMID", ""), r.get("DOI", "")
        title = r.get("Title", "")
        print(f"      Ref[{r.get('Ref Index')}] PMID:{pmid or 'N/A'} ...", end=" ", flush=True)
        url = fetcher.find_pdf_url(pmid, doi)
        if not url:
            print("无 PDF")
            r["PDF_Reason"] = "未找到开放获取 PDF"
            continue
        r["PDF_URL"] = url
        path = fetcher.download(url, pmid or f"ref{r.get('Ref Index')}", subdir="refs")
        if not path:
            print("下载失败")
            r["PDF_Reason"] = "PDF 下载失败"
            continue
        r["PDF_File"] = path
        text = fetcher.extract_text(path, analyzer.pdf_text_max_len)
        if not text:
            print("无文本")
            r["PDF_Reason"] = "PDF 文本提取为空"
            continue
        review = analyzer.review_ref_pdf(pmid, title, text)
        rel = review.get("is_relevant")
        r["PDF_Relevant"] = "" if rel is None else bool(rel)
        r["PDF_Confidence"] = review.get("confidence", "")
        r["PDF_Reason"] = review.get("reason", "")
        print(f"复查={rel}")
        time.sleep(sleep_sec)


# ---------------------------------------------------------------------------
# 单篇文献处理 + 导出
# ---------------------------------------------------------------------------

SOURCE_LABELS = {
    "pdf": "PDF 全文",
    "pmc_web": "PMC 在线全文(XML)",
    "pmc_html": "PMC 网页全文",
    "abstract": "仅摘要(版本1逻辑)",
}


def process_article(pmid: str, cfg: Dict, ncbi: NCBIClient,
                    analyzer: ArticleAnalyzer,
                    fetcher: Optional[PDFFetcher]) -> Tuple[Dict, List[Dict]]:
    """处理单篇检索命中文献，返回 (文献记录, 参考文献记录列表)。"""
    print(f"\n{'='*70}\n  处理 PMID: {pmid}\n{'='*70}")

    ac = acquire_content(pmid, ncbi, fetcher, cfg)
    meta = ac.meta
    title = meta.get("Title", "")
    abstract = meta.get("Abstract", "")
    result_key = cfg.get("result_key", "is_relevant")

    article_rec = {
        "PMID": pmid,
        "Title": title,
        "Abstract": abstract,
        "Content Source": SOURCE_LABELS.get(ac.source, ac.source),
        "PMCID": ac.pmcid,
        "PDF_URL": ac.pdf_url,
        "PDF_File": ac.pdf_file,
        "Article_Relevant": "",
        "Article_Confidence": "",
        "Article_Reason": "",
        "Has_Potential_Refs": "",
        "Potential_Refs_Count": 0,
    }

    ref_records: List[Dict] = []

    if ac.source == "abstract":
        print("  [AI] 基于摘要判断（版本 1 逻辑）...")
        judgment = analyzer.judge_article_abstract(pmid, title, abstract)
        rel = judgment.get("is_relevant")
        article_rec["Article_Relevant"] = "" if rel is None else bool(rel)
        article_rec["Article_Confidence"] = judgment.get("confidence", "")
        article_rec["Article_Reason"] = judgment.get("reason", "")
        article_rec["Has_Potential_Refs"] = "N/A"
        print(f"  结果: relevant={rel}")
        time.sleep(cfg.get("ai_sleep_sec", 1.0))
        return article_rec, ref_records

    # 有全文（PDF / PMC）
    print(f"  [AI] 基于{SOURCE_LABELS.get(ac.source, ac.source)}判断文献...")
    judgment = analyzer.judge_article_fulltext(pmid, title, ac.fulltext)
    rel = judgment.get("is_relevant")
    article_rec["Article_Relevant"] = "" if rel is None else bool(rel)
    article_rec["Article_Confidence"] = judgment.get("confidence", "")
    article_rec["Article_Reason"] = judgment.get("reason", "")
    print(f"  文献判断: relevant={rel}")
    time.sleep(cfg.get("ai_sleep_sec", 1.0))

    # 参考文献分析
    if ac.refs:
        print(f"  [AI] 结构化参考文献筛选 ({len(ac.refs)} 条)...")
        if cfg.get("enrich_references", True):
            enrich_references(ac.refs, ncbi, cfg.get("ref_enrich_batch_size", 100))
        kept = screen_structured_refs(
            analyzer, ac.fulltext, ac.refs,
            cfg.get("ai_batch_size", 10), cfg.get("ai_sleep_sec", 1.0),
            result_key,
        )
        for r in kept:
            r["Source PMID"] = pmid
        if fetcher and kept and cfg.get("download_pdf", True):
            print(f"  [PDF] 对 {len(kept)} 条命中文献下载 PDF 复查...")
            pdf_review_refs(kept, fetcher, analyzer, cfg.get("ai_sleep_sec", 1.0))
        ref_records = kept
        article_rec["Has_Potential_Refs"] = len(kept) > 0
        article_rec["Potential_Refs_Count"] = len(kept)
        print(f"  命中参考文献: {len(kept)}/{len(ac.refs)}")

    elif ac.source == "pdf":
        print("  [AI] 从 PDF 全文挖掘参考文献...")
        mined = analyzer.find_refs_in_pdf_text(pmid, title, ac.fulltext)
        for i, r in enumerate(mined, 1):
            if r.get("is_potentially_relevant"):
                ref_records.append({
                    "Source PMID": pmid,
                    "Ref Index": i,
                    "PMID": "",
                    "DOI": "",
                    "Title": r.get("citation", ""),
                    "Authors": "",
                    "Source": "",
                    "Year": "",
                    "Abstract": "",
                    "AI_Confidence": r.get("confidence", ""),
                    "AI_Reason": r.get("reason", ""),
                    "Ref_Info": r.get("info", ""),
                    "Source Type": "pdf_text_mining",
                    "PDF_URL": "", "PDF_File": "",
                    "PDF_Relevant": "", "PDF_Confidence": "", "PDF_Reason": "",
                })
        article_rec["Has_Potential_Refs"] = len(ref_records) > 0
        article_rec["Potential_Refs_Count"] = len(ref_records)
        print(f"  挖掘到可能相关参考文献: {len(ref_records)} 条")
        time.sleep(cfg.get("ai_sleep_sec", 1.0))
    else:
        article_rec["Has_Potential_Refs"] = False
        article_rec["Potential_Refs_Count"] = 0
        print("  ℹ️  无结构化参考文献且非 PDF 来源，跳过参考文献分析")

    return article_rec, ref_records


def _style_worksheet(ws, row_height: int = 50):
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    even = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = even if row_idx % 2 == 0 else odd
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[row_idx].height = row_height


def export_results(article_records: List[Dict], ref_records: List[Dict], out_file: str):
    article_cols = [
        "PMID", "Title", "Abstract", "Content Source", "PMCID",
        "PDF_URL", "PDF_File",
        "Article_Relevant", "Article_Confidence", "Article_Reason",
        "Has_Potential_Refs", "Potential_Refs_Count",
    ]
    ref_cols = [
        "Source PMID", "Ref Index", "PMID", "DOI", "Title", "Authors",
        "Source", "Year", "Abstract", "Ref_Info",
        "AI_Confidence", "AI_Reason", "Source Type",
        "PDF_URL", "PDF_File", "PDF_Relevant", "PDF_Confidence", "PDF_Reason",
    ]

    os.makedirs(os.path.dirname(out_file) or ".", exist_ok=True)
    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        adf = pd.DataFrame(article_records)
        adf = adf[[c for c in article_cols if c in adf.columns]]
        adf.reset_index(drop=True, inplace=True)
        adf.index += 1
        adf.to_excel(writer, index=True, index_label="No.", sheet_name="Articles")

        if ref_records:
            rdf = pd.DataFrame(ref_records)
            rdf = rdf[[c for c in ref_cols if c in rdf.columns]]
            rdf.reset_index(drop=True, inplace=True)
            rdf.index += 1
            rdf.to_excel(writer, index=True, index_label="No.", sheet_name="References")
        else:
            pd.DataFrame(columns=ref_cols).to_excel(
                writer, index=False, sheet_name="References")

        for sheet in writer.sheets.values():
            _style_worksheet(sheet, row_height=60 if sheet.title == "References" else 50)

    rel_count = sum(1 for r in article_records if r.get("Article_Relevant") is True)
    print(f"\n{'='*70}")
    print(f"  完成！")
    print(f"  文件: {out_file}")
    print(f"  检索文献: {len(article_records)} 篇，符合条件: {rel_count} 篇")
    print(f"  参考文献记录: {len(ref_records)} 条")
    print(f"{'='*70}\n")


def main():
    print("=" * 70)
    print("  PubMed 全文版（V2）：检索式驱动 + 全文/摘要分级分析")
    print("=" * 70)

    cfg = load_config("config_v2.json")

    query = cfg.get("query", "").strip()
    seeds = [str(s) for s in cfg.get("seed_pmids", [])]
    if not query and not seeds:
        print("\n请在 config_v2.json 中填写 query（检索式）或 seed_pmids")
        return

    ai_key = os.getenv("AI_API_KEY") or cfg.get("ai_api_key")
    if not ai_key or ai_key == "your_api_key_here":
        print("\n未设置 AI 密钥，请设置环境变量 AI_API_KEY 或填写 config_v2.json 的 ai_api_key")
        return

    ncbi = NCBIClient(cfg.get("ncbi_api_key"), cfg.get("ncbi_sleep_sec", 0.4))
    analyzer = ArticleAnalyzer(cfg, ai_key)

    fetcher = None
    if cfg.get("download_pdf", True):
        if PdfReader is None:
            print("\n未安装 pypdf，PDF 下载/提取已关闭（pip install pypdf）")
        else:
            fetcher = PDFFetcher(
                ncbi, cfg.get("pdf_dir", "output-v2/pdfs"),
                cfg.get("unpaywall_email"),
                cfg.get("pdf_download_timeout", 60),
            )

    pmids = search_pmids(cfg, ncbi) if query else []
    all_pmids = list(dict.fromkeys(pmids + seeds))
    if not all_pmids:
        print("\n未获取到任何 PMID")
        return
    print(f"\n待处理文献: {len(all_pmids)} 篇")

    all_articles: List[Dict] = []
    all_refs: List[Dict] = []
    for pmid in all_pmids:
        try:
            art, refs = process_article(pmid, cfg, ncbi, analyzer, fetcher)
            all_articles.append(art)
            all_refs.extend(refs)
        except (requests.RequestException, ET.ParseError) as e:
            print(f"  PMID {pmid} 处理失败: {e}")

    if not all_articles:
        print("\n没有处理任何文献")
        return

    export_results(all_articles, all_refs,
                   cfg.get("out_file", "output-v2/screening_results.xlsx"))


if __name__ == "__main__":
    main()
