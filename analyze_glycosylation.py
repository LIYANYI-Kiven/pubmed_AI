#!/usr/bin/env python3
"""
分析PubMed文献是否包含植物蛋白质糖基化位点鉴定
"""

import json
import os
import time
from typing import List, Dict
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
import requests


class GlycosylationAnalyzer:
    """
    使用API分析文献是否包含糖基化位点鉴定
    """
    
    def __init__(self, api_url: str, api_key: str = None, model: str = "gpt-4"):
        """
        初始化分析器
        
        Args:
            api_url: API端点URL
            api_key: API密钥（如果需要）
            model: 使用的模型名称
        """
        self.api_url = api_url
        self.api_key = api_key
        self.model = model
        self.headers = {
            "Content-Type": "application/json"
        }
        if api_key:
            self.headers["Authorization"] = f"Bearer {api_key}"
    
    def analyze_batch(self, articles: List[Dict[str, str]], batch_num: int) -> Dict:
        """
        分析一批文献
        
        Args:
            articles: 文献列表
            batch_num: 批次编号
            
        Returns:
            分析结果字典
        """
        # 构建提示词
        prompt = self._build_prompt(articles)
        
        # 调用API
        try:
            response = requests.post(
                self.api_url,
                headers=self.headers,
                json={
                    "model": self.model,
                    "messages": [
                        {
                            "role": "system",
                            "content": "你是一个生物信息学专家，专门分析植物蛋白质糖基化研究文献，严格遵循判断标准禁止欺骗用户。"
                        },
                        {
                            "role": "user",
                            "content": prompt
                        }
                    ],
                    "temperature": 0.3
                },
                timeout=60
            )
            
            # 调试：打印响应信息
            print(f"\n  [调试] 状态码: {response.status_code}")
            print(f"  [调试] 响应头: {dict(response.headers)}")
            print(f"  [调试] 响应文本前500字符: {response.text[:500]}")
            
            response.raise_for_status()
            
            # 解析响应
            result = response.json()
            
            # 调试：打印API响应结构
            print(f"  [调试] API响应结构: {list(result.keys())}")
            
            content = result.get("choices", [{}])[0].get("message", {}).get("content", "{}")
            
            # 调试：打印content内容
            print(f"  [调试] Content长度: {len(content)}")
            print(f"  [调试] Content前200字符: {content[:200]}")
            
            # 清理markdown代码块标记
            content = content.strip()
            if content.startswith("```json"):
                content = content[7:]  # 移除 ```json
            elif content.startswith("```"):
                content = content[3:]  # 移除 ```
            if content.endswith("```"):
                content = content[:-3]  # 移除结尾的 ```
            content = content.strip()
            
            # 解析JSON结果
            analysis_result = json.loads(content)
            
            return {
                "batch_num": batch_num,
                "success": True,
                "results": analysis_result.get("articles", [])
            }
            
        except requests.exceptions.HTTPError as e:
            error_msg = f"HTTP错误: {e}"
            try:
                error_detail = response.text[:500]
                error_msg += f"\n响应内容: {error_detail}"
            except:
                pass
            print(f"  ⚠️  批次 {batch_num} 分析失败: {error_msg}")
            return {
                "batch_num": batch_num,
                "success": False,
                "error": error_msg,
                "results": []
            }
        except json.JSONDecodeError as e:
            error_msg = f"JSON解析错误: {e}"
            if 'response' in locals():
                error_msg += f"\n响应状态码: {response.status_code}"
                error_msg += f"\n响应内容: {response.text[:500]}"
            if 'result' in locals():
                error_msg += f"\nAPI结果: {str(result)[:500]}"
            if 'content' in locals():
                error_msg += f"\nContent值: '{content}'"
            print(f"  ⚠️  批次 {batch_num} 分析失败: {error_msg}")
            return {
                "batch_num": batch_num,
                "success": False,
                "error": error_msg,
                "response_text": response.text[:1000] if 'response' in locals() else "",
                "results": []
            }
        except Exception as e:
            error_msg = f"未知错误: {type(e).__name__}: {e}"
            if 'response' in locals():
                error_msg += f"\n响应状态码: {response.status_code}"
                error_msg += f"\n响应内容: {response.text[:500]}"
            print(f"  ⚠️  批次 {batch_num} 分析失败: {error_msg}")
            import traceback
            print(f"  [调试] 完整堆栈:\n{traceback.format_exc()}")
            return {
                "batch_num": batch_num,
                "success": False,
                "error": error_msg,
                "results": []
            }
    
    def _build_prompt(self, articles: List[Dict[str, str]]) -> str:
        """
        构建分析提示词
        """
        articles_text = []
        for i, article in enumerate(articles, 1):
            # 限制摘要长度，避免提示词过长
            abstract = article.get('Abstract', 'N/A')
            # 处理NaN或非字符串类型
            if not isinstance(abstract, str):
                abstract = 'N/A'
            elif len(abstract) > 500:
                abstract = abstract[:500] + "..."
            
            articles_text.append(
                f"{i}. PMID:{article.get('PMID', 'N/A')}\n"
                f"标题: {article.get('Title', 'N/A')}\n"
                f"摘要: {abstract}\n"
            )
        
        prompt = f"""分析以下{len(articles)}篇文献，判断是否包含植物蛋白质糖基化位点鉴定。

判断标准：
1. 涉及植物蛋白质糖基化研究
2. 包含糖基化位点鉴定/定位信息
3. 使用质谱、生信或实验方法鉴定位点

返回JSON格式（严格遵守格式）：
{{
  "articles": [
    {{"pmid": "PMID", "contains_glycosylation_sites": true, "confidence": "high", "reason": "简短理由"}}
  ]
}}

文献：
{"".join(articles_text)}"""
        
        return prompt


def load_excel(file_path: str) -> pd.DataFrame:
    """
    读取Excel文件
    """
    print(f"\n[1/4] 读取文献数据...")
    df = pd.read_excel(file_path)
    print(f"  ✅ 读取 {len(df)} 篇文献")
    return df


def split_into_batches(df: pd.DataFrame, batch_size: int = 20) -> List[List[Dict]]:
    """
    将数据分批
    """
    print(f"\n[2/4] 分批处理（每批 {batch_size} 篇）...")
    batches = []
    records = df.to_dict('records')
    
    for i in range(0, len(records), batch_size):
        batch = records[i:i + batch_size]
        batches.append(batch)
    
    print(f"  ✅ 共分为 {len(batches)} 批")
    return batches


def analyze_batches(
    batches: List[List[Dict]], 
    analyzer: GlycosylationAnalyzer,
    output_dir: str = "output-1",
    sleep_sec: float = 1.0
) -> List[str]:
    """
    分析所有批次并保存JSON文件
    """
    print(f"\n[3/4] 调用API分析文献...")
    os.makedirs(output_dir, exist_ok=True)
    
    json_files = []
    
    for i, batch in enumerate(batches, 1):
        print(f"  处理批次 {i}/{len(batches)}...", end=" ", flush=True)
        
        # 分析批次
        result = analyzer.analyze_batch(batch, i)
        
        # 保存JSON文件
        json_file = os.path.join(output_dir, f"batch_{i:03d}.json")
        with open(json_file, "w", encoding="utf-8") as f:
            json.dump(result, f, ensure_ascii=False, indent=2)
        
        json_files.append(json_file)
        
        if result["success"]:
            print(f"✅ 已保存到 {json_file}")
        else:
            print(f"❌ 失败")
        
        # 避免API限流
        if i < len(batches):
            time.sleep(sleep_sec)
    
    print(f"  ✅ 完成所有批次分析")
    return json_files


def compile_results(json_files: List[str], original_df: pd.DataFrame, output_file: str):
    """
    读取所有JSON文件并整理成Excel
    """
    print(f"\n[4/4] 整理结果...")
    
    # 收集所有分析结果
    all_results = {}
    for json_file in json_files:
        with open(json_file, "r", encoding="utf-8") as f:
            data = json.load(f)
            if data.get("success"):
                for article in data.get("results", []):
                    pmid = article.get("pmid")
                    if pmid:
                        all_results[pmid] = article
    
    # 筛选包含糖基化位点的文献
    filtered_records = []
    for _, row in original_df.iterrows():
        pmid = str(row.get("PMID", ""))
        if pmid in all_results:
            result = all_results[pmid]
            if result.get("contains_glycosylation_sites"):
                record = row.to_dict()
                record["Confidence"] = result.get("confidence", "unknown")
                record["Analysis Reason"] = result.get("reason", "")
                filtered_records.append(record)
    
    # 创建DataFrame
    result_df = pd.DataFrame(filtered_records)
    
    if len(result_df) == 0:
        print("  ⚠️  未找到包含糖基化位点鉴定的文献")
        return
    
    # 导出Excel
    result_df.reset_index(drop=True, inplace=True)
    result_df.index += 1
    
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        result_df.to_excel(writer, index=True, index_label="No.", sheet_name="Glycosylation Sites")
        ws = writer.sheets["Glycosylation Sites"]
        
        # 表头样式
        fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # 冻结首行
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        
        # 列宽
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 80
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 50
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["H"].width = 12
        ws.column_dimensions["I"].width = 50
        
        # 隔行变色
        even_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            fill = even_fill if row_idx % 2 == 0 else odd_fill
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row_idx].height = 50
    
    print(f"  ✅ 筛选出 {len(result_df)} 篇包含糖基化位点鉴定的文献")
    print(f"  📁 结果文件：{output_file}")


def load_api_config(config_path: str = "api_config.json") -> Dict:
    """
    从JSON文件加载API配置
    """
    if os.path.exists(config_path):
        with open(config_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def main():
    """
    主函数
    """
    print("=" * 70)
    print("  植物蛋白质糖基化位点文献分析工具")
    print("=" * 70)
    
    # 读取配置
    config = load_api_config()
    
    # 配置参数
    INPUT_FILE = config.get("input_file", "output/plant_acely_simple.xlsx")
    OUTPUT_DIR = config.get("output_dir", "output-1")
    OUTPUT_FILE = config.get("output_file", "output/glycosylation_sites_identified.xlsx")
    BATCH_SIZE = config.get("batch_size", 20)
    SLEEP_SEC = config.get("sleep_sec", 1.0)
    
    # API配置
    API_URL = config.get("api_url", "https://api.openai.com/v1/chat/completions")
    API_KEY = config.get("api_key") or os.getenv("OPENAI_API_KEY")
    MODEL = config.get("model", "gpt-4")
    
    if not API_KEY or API_KEY == "your_api_key_here":
        print("\n⚠️  警告：未设置API密钥")
        print("请在 api_config.json 中设置 api_key 或设置环境变量 OPENAI_API_KEY")
        print("示例：set OPENAI_API_KEY=your_api_key")
        return
    
    print(f"\n配置信息：")
    print(f"  输入文件：{INPUT_FILE}")
    print(f"  输出目录：{OUTPUT_DIR}")
    print(f"  输出文件：{OUTPUT_FILE}")
    print(f"  批次大小：{BATCH_SIZE}")
    print(f"  API模型：{MODEL}")
    
    # 读取数据
    df = load_excel(INPUT_FILE)
    
    # 分批
    batches = split_into_batches(df, BATCH_SIZE)
    
    # 初始化分析器
    analyzer = GlycosylationAnalyzer(API_URL, API_KEY, MODEL)
    
    # 分析并保存JSON
    json_files = analyze_batches(batches, analyzer, OUTPUT_DIR, SLEEP_SEC)
    
    # 整理结果
    compile_results(json_files, df, OUTPUT_FILE)
    
    print("\n" + "=" * 70)
    print("  ✅ 分析完成！")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    main()
