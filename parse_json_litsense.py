#!/usr/bin/env python3
"""
parse_json_litsense.py

整理 litsense_search.py 生成的 JSON 文件（句子或段落模式）。
LitSense 返回结果中只有 pmid/pmcid，没有标题和 DOI，
本脚本通过 PubMed E-utils 批量回查补全这些信息。

提取字段：标题、PubMed ID、DOI、PubMed 摘要链接。

用法:
    python parse_json_litsense.py
    python parse_json_litsense.py --input output/litsense_raw.json --output output/litsense_summary.xlsx
"""

import argparse
import json
import os
import time
from typing import Dict, List, Optional
import xml.etree.ElementTree as ET

import requests
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


# ---------------------------------------------------------------------------
# PubMed 回查（获取标题和 DOI）
# ---------------------------------------------------------------------------

def fetch_pubmed_meta(pmids: List[str], api_key: Optional[str] = None,
                      batch_size: int = 200, sleep_sec: float = 0.5) -> Dict[str, Dict]:
    """
    通过 PubMed efetch 批量获取标题和 DOI。
    返回 {pmid: {"title": ..., "doi": ...}}
    """
    if not pmids:
        return {}

    base = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/"
    params_base = {"api_key": api_key} if api_key else {}
    session = requests.Session()
    session.headers["User-Agent"] = "parse-litsense/1.0"

    result = {}
    for i in range(0, len(pmids), batch_size):
        chunk = pmids[i:i + batch_size]
        try:
            r = session.get(
                base + "efetch.fcgi",
                params={
                    "db": "pubmed", "id": ",".join(chunk),
                    "retmode": "xml", "rettype": "abstract",
                    **params_base,
                },
                timeout=60,
            )
            r.raise_for_status()
            root = ET.fromstring(r.content)
            for article in root.findall(".//PubmedArticle"):
                pmid = article.findtext(".//PMID", "").strip()
                title_elem = article.find(".//ArticleTitle")
                title = "".join(title_elem.itertext()).strip() if title_elem is not None else ""
                doi = ""
                for eid in article.findall(".//ArticleId"):
                    if eid.attrib.get("IdType") == "doi":
                        doi = (eid.text or "").strip()
                        break
                if pmid:
                    result[pmid] = {"title": title, "doi": doi}
        except Exception as e:
            print(f"  ⚠️  PubMed 回查失败 (batch {i}): {e}")
        time.sleep(sleep_sec)

    return result


# ---------------------------------------------------------------------------
# 解析 LitSense JSON
# ---------------------------------------------------------------------------

def parse_litsense_json(input_file: str, api_key: Optional[str] = None) -> List[Dict]:
    """
    从 litsense_search.py 生成的 JSON 中提取条目，
    去重后回查 PubMed 补全标题/DOI，返回汇总列表。
    """
    with open(input_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    # LitSense JSON 结构：{"sentence": [...]} 或 {"passage": [...]} 或直接列表
    if isinstance(data, list):
        items = data
    else:
        # 取第一个 key 下的列表（sentence / passage）
        items = next(iter(data.values()), []) if data else []

    # 去重：同一 pmid 可能出现多次（多个匹配句子）
    seen_pmids: Dict[str, Dict] = {}  # pmid -> 代表条目
    for item in items:
        pmid  = str(item.get("pmid") or "")
        pmcid = str(item.get("pmcid") or "")
        if not pmid:
            continue
        if pmid not in seen_pmids:
            seen_pmids[pmid] = {
                "pmid":  pmid,
                "pmcid": pmcid if pmcid != "None" else "",
            }

    print(f"  发现 {len(items)} 条结果，去重后 {len(seen_pmids)} 篇文献")

    if not seen_pmids:
        return []

    # 回查 PubMed 补全标题和 DOI
    print("  回查 PubMed 补全标题和 DOI...")
    meta_map = fetch_pubmed_meta(list(seen_pmids.keys()), api_key=api_key)

    records = []
    for i, (pmid, info) in enumerate(seen_pmids.items(), 1):
        meta  = meta_map.get(pmid, {})
        doi   = meta.get("doi", "")
        title = meta.get("title", "")

        pubmed_url = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/" if pmid else ""
        doi_url    = f"https://doi.org/{doi}" if doi else ""

        records.append({
            "No.":             i,
            "Title":           title,
            "PubMed ID":       pmid,
            "PMCID":           info.get("pmcid", ""),
            "DOI":             doi,
            "PubMed Abstract": pubmed_url,
            "DOI URL":         doi_url,
        })

    return records


# ---------------------------------------------------------------------------
# 导出
# ---------------------------------------------------------------------------

def export_excel(records: List[Dict], out_file: str):
    df = pd.DataFrame(records).set_index("No.")
    os.makedirs(os.path.dirname(out_file) if os.path.dirname(out_file) else ".", exist_ok=True)

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Summary")
        ws = writer.sheets["Summary"]

        fill = PatternFill(start_color="2E6B9E", end_color="2E6B9E", fill_type="solid")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        col_widths = {"A": 6, "B": 70, "C": 12, "D": 16, "E": 35, "F": 45, "G": 45}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        even = PatternFill(start_color="EAF2FB", end_color="EAF2FB", fill_type="solid")
        odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.fill = even if row_idx % 2 == 0 else odd
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row_idx].height = 40

    print(f"  📊 Excel: {out_file}（{len(df)} 篇）")


# ---------------------------------------------------------------------------
# 主函数
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="整理 LitSense JSON，补全标题和 DOI")
    parser.add_argument("--input",   default="output/litsense_raw.json",      help="输入 JSON 路径")
    parser.add_argument("--output",  default="output/litsense_summary.xlsx",  help="输出 Excel 路径")
    parser.add_argument("--api-key", default=None, help="NCBI API Key（可选）")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"❌ 文件不存在: {args.input}")
        return

    print(f"读取: {args.input}")
    records = parse_litsense_json(args.input, api_key=args.api_key)

    if not records:
        print("⚠️  无有效记录")
        return

    print(f"共 {len(records)} 篇唯一文献")
    export_excel(records, args.output)

    out_json = os.path.splitext(args.output)[0] + "_detail.json"
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    print(f"  📄 JSON:  {out_json}")


if __name__ == "__main__":
    main()
