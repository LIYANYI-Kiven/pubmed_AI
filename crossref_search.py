#!/usr/bin/env python3
"""
Crossref 文献检索工具

Crossref REST API 提供超过 1.5 亿篇学术文献的元数据检索，
涵盖 DOI、标题、作者、摘要、期刊、引用次数等信息。

API 文档：https://api.crossref.org
无需注册，建议在 User-Agent 或 mailto 参数中提供联系邮箱
以获得更好的服务质量（polite pool）。

用法：
  python crossref_search.py --query "plant protein glycosylation" --rows 50
  python crossref_search.py --config crossref_config.json
"""

import argparse
import json
import os
import time
from typing import List, Dict, Optional
from urllib.parse import quote

import requests
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


# ---------------------------------------------------------------------------
# Crossref API 客户端
# ---------------------------------------------------------------------------

class CrossrefClient:
    """
    Crossref REST API 客户端。

    API 特点：
      - 免费，无需 API Key
      - 提供 mailto 参数（或 User-Agent）可进入 polite pool，速率更稳定
      - 支持全文检索、字段过滤、日期过滤、结果排序
      - 单次最多返回 1000 条结果（rows 参数）
      - 支持游标（cursor）分页，可获取全量结果
    """

    BASE_URL = "https://api.crossref.org/works"

    def __init__(self, mailto: Optional[str] = None, sleep_sec: float = 1.0, timeout: float = 30.0):
        """
        Args:
            mailto:    你的联系邮箱（强烈建议提供，加入 polite pool 享受更稳定服务）
            sleep_sec: 请求间隔（秒）
            timeout:   单次请求超时（秒）
        """
        self.mailto = mailto
        self.sleep_sec = sleep_sec
        self.timeout = timeout
        self.session = requests.Session()
        ua = "crossref-search/1.0"
        if mailto:
            ua += f" (mailto:{mailto})"
        self.session.headers.update({"User-Agent": ua})

    def search(
        self,
        query: str,
        rows: int = 20,
        offset: int = 0,
        sort: str = "relevance",
        order: str = "desc",
        filter_from_year: Optional[int] = None,
        filter_to_year: Optional[int] = None,
        filter_type: Optional[str] = None,
        filter_has_abstract: bool = False,
        select_fields: Optional[List[str]] = None,
    ) -> Dict:
        """
        执行文献检索。

        Args:
            query:             检索关键词（自由文本）
            rows:              返回结果数量（最大 1000）
            offset:            结果偏移量（用于分页）
            sort:              排序方式："relevance"、"published"、"is-referenced-by-count"
            order:             排序顺序："asc" 或 "desc"
            filter_from_year:  发表年份起始（如 2017）
            filter_to_year:    发表年份截止（如 2025）
            filter_type:       文献类型过滤，如 "journal-article"、"proceedings-article"
            filter_has_abstract: 是否只返回有摘要的文献
            select_fields:     指定返回字段列表（None 返回所有字段）

        Returns:
            API 返回的完整 JSON 字典，含 message.items 列表
        """
        params: Dict = {
            "query": query,
            "rows":  min(rows, 1000),
            "offset": offset,
            "sort":  sort,
            "order": order,
        }

        if self.mailto:
            params["mailto"] = self.mailto

        # 构建 filter 字符串
        filters = []
        if filter_from_year:
            filters.append(f"from-pub-date:{filter_from_year}")
        if filter_to_year:
            filters.append(f"until-pub-date:{filter_to_year}")
        if filter_type:
            filters.append(f"type:{filter_type}")
        if filter_has_abstract:
            filters.append("has-abstract:true")
        if filters:
            params["filter"] = ",".join(filters)

        if select_fields:
            params["select"] = ",".join(select_fields)

        r = self.session.get(self.BASE_URL, params=params, timeout=self.timeout)
        r.raise_for_status()
        return r.json()

    def search_all(
        self,
        query: str,
        max_results: int = 100,
        batch_size: int = 100,
        **kwargs,
    ) -> List[Dict]:
        """
        分页获取全部结果，直到达到 max_results 或无更多结果。

        Args:
            query:       检索关键词
            max_results: 最大获取数量
            batch_size:  每次 API 请求的数量
            **kwargs:    其他参数传递给 search()

        Returns:
            所有文献记录的列表
        """
        all_items = []
        offset = 0

        while len(all_items) < max_results:
            remaining = max_results - len(all_items)
            rows = min(batch_size, remaining, 1000)

            try:
                data = self.search(query=query, rows=rows, offset=offset, **kwargs)
                items = data.get("message", {}).get("items", [])
                total_results = data.get("message", {}).get("total-results", 0)

                if not items:
                    break

                all_items.extend(items)
                offset += len(items)

                print(f"  进度: {len(all_items)}/{min(max_results, total_results)} "
                      f"（数据库总计 {total_results} 条）", end="\r", flush=True)

                if offset >= total_results:
                    break

                time.sleep(self.sleep_sec)

            except Exception as e:
                print(f"\n  ⚠️  批次 offset={offset} 失败: {e}")
                break

        print()
        return all_items

    def get_by_doi(self, doi: str) -> Optional[Dict]:
        """通过 DOI 获取单篇文献元数据。"""
        url = f"https://api.crossref.org/works/{quote(doi, safe='')}"
        try:
            r = self.session.get(url, timeout=self.timeout)
            r.raise_for_status()
            return r.json().get("message")
        except Exception:
            return None

    def get_journal_info(self, issn: str) -> Optional[Dict]:
        """通过 ISSN 获取期刊信息。"""
        url = f"https://api.crossref.org/journals/{issn}"
        try:
            r = self.session.get(url, timeout=self.timeout)
            r.raise_for_status()
            return r.json().get("message")
        except Exception:
            return None


# ---------------------------------------------------------------------------
# 结果解析
# ---------------------------------------------------------------------------

def parse_item(item: Dict) -> Dict:
    """将单条 Crossref API 结果规范化为统一字段。"""

    # DOI
    doi = item.get("DOI", "")

    # 标题
    titles = item.get("title", [])
    title = titles[0] if titles else ""

    # 作者
    authors = item.get("author", [])
    author_names = []
    for a in authors:
        name = a.get("family", "")
        given = a.get("given", "")
        if given:
            name = f"{name} {given}"
        if name:
            author_names.append(name)
    first_author = author_names[0] if author_names else ""
    all_authors = "; ".join(author_names[:10]) + (" et al." if len(author_names) > 10 else "")

    # 期刊
    container = item.get("container-title", [])
    journal = container[0] if container else ""
    short_journal = item.get("short-container-title", [""])[0] if item.get("short-container-title") else ""

    # 发表日期
    pub_date = ""
    for date_key in ["published", "published-print", "published-online", "created"]:
        dp = item.get(date_key, {}).get("date-parts", [[]])
        if dp and dp[0]:
            parts = dp[0]
            if len(parts) >= 3:
                pub_date = f"{parts[0]}-{parts[1]:02d}-{parts[2]:02d}"
            elif len(parts) == 2:
                pub_date = f"{parts[0]}-{parts[1]:02d}"
            elif len(parts) == 1:
                pub_date = str(parts[0])
            if pub_date:
                break

    # 摘要
    abstract = item.get("abstract", "")
    # 清理 JATS XML 标签
    if abstract:
        import re
        abstract = re.sub(r"<[^>]+>", " ", abstract)
        abstract = re.sub(r"\s+", " ", abstract).strip()

    # 引用次数
    cited_by = item.get("is-referenced-by-count", 0)

    # 其他信息
    item_type = item.get("type", "")
    publisher = item.get("publisher", "")
    issn_list = item.get("ISSN", [])
    issn = issn_list[0] if issn_list else ""
    volume = item.get("volume", "")
    issue = item.get("issue", "")
    pages = item.get("page", "")
    url = f"https://doi.org/{doi}" if doi else ""

    # 关联 PMID（如果 Crossref 提供了）
    pmid = ""
    for link in item.get("link", []):
        if "pubmed" in link.get("URL", "").lower():
            pmid = link["URL"].split("/")[-1]
            break

    return {
        "DOI":           doi,
        "Title":         title,
        "First Author":  first_author,
        "Authors":       all_authors,
        "Journal":       journal,
        "Short Journal": short_journal,
        "ISSN":          issn,
        "Publisher":     publisher,
        "Volume":        volume,
        "Issue":         issue,
        "Pages":         pages,
        "Published":     pub_date,
        "Type":          item_type,
        "Cited By":      cited_by,
        "Abstract":      abstract,
        "PMID":          pmid,
        "URL":           url,
    }


def export_excel(records: List[Dict], out_file: str, query_info: str = ""):
    """导出格式化 Excel。"""
    if not records:
        print("  ⚠️  无结果可导出")
        return

    df = pd.DataFrame(records)

    # 按引用次数降序排序
    if "Cited By" in df.columns:
        df = df.sort_values("Cited By", ascending=False)
    df.reset_index(drop=True, inplace=True)
    df.index += 1

    os.makedirs(os.path.dirname(out_file) if os.path.dirname(out_file) else ".", exist_ok=True)

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=True, index_label="No.", sheet_name="Crossref Results")
        ws = writer.sheets["Crossref Results"]

        header_fill = PatternFill(start_color="1B6CA8", end_color="1B6CA8", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions

        col_widths = {
            "A": 6,  "B": 35, "C": 60, "D": 20, "E": 40,
            "F": 25, "G": 20, "H": 10, "I": 20, "J": 8,
            "K": 8,  "L": 8,  "M": 14, "N": 15, "O": 8,
            "P": 80, "Q": 12, "R": 40,
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        even_fill = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")
        odd_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.fill = even_fill if row_idx % 2 == 0 else odd_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row_idx].height = 50

    print(f"  📁 已导出: {out_file}（{len(df)} 篇文献）")


def export_json(records: List[Dict], raw_items: List[Dict], out_file: str):
    """保存原始 API 结果。"""
    os.makedirs(os.path.dirname(out_file) if os.path.dirname(out_file) else ".", exist_ok=True)
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(raw_items, f, ensure_ascii=False, indent=2)
    print(f"  📁 原始结果: {out_file}")


# ---------------------------------------------------------------------------
# 配置加载
# ---------------------------------------------------------------------------

DEFAULT_CONFIG = {
    "query":               "",
    "max_results":         100,
    "batch_size":          100,
    "sort":                "relevance",
    "order":               "desc",
    "filter_from_year":    None,
    "filter_to_year":      None,
    "filter_type":         "journal-article",
    "filter_has_abstract": False,
    "mailto":              None,
    "sleep_sec":           1.0,
    "timeout":             30.0,
    "out_excel":           "output/crossref_results.xlsx",
    "out_json":            "output/crossref_raw.json",
}


# ---------------------------------------------------------------------------
# 主函数
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Crossref 文献检索工具")
    parser.add_argument("--query",  type=str, help="检索关键词")
    parser.add_argument("--rows",   type=int, default=100, help="最大结果数量（默认 100）")
    parser.add_argument("--from",   type=int, dest="from_year", help="发表年份起始")
    parser.add_argument("--to",     type=int, dest="to_year",   help="发表年份截止")
    parser.add_argument("--sort",   type=str, default="relevance",
                        choices=["relevance", "published", "is-referenced-by-count"],
                        help="排序方式")
    parser.add_argument("--mailto", type=str, help="联系邮箱（建议提供）")
    parser.add_argument("--config", type=str, help="JSON 配置文件路径")
    parser.add_argument("--out",    type=str, help="输出 Excel 文件路径")
    args = parser.parse_args()

    config = dict(DEFAULT_CONFIG)
    if args.config and os.path.exists(args.config):
        with open(args.config, "r", encoding="utf-8") as f:
            config.update(json.load(f))

    if args.query:
        config["query"] = args.query
    if args.rows:
        config["max_results"] = args.rows
    if args.from_year:
        config["filter_from_year"] = args.from_year
    if args.to_year:
        config["filter_to_year"] = args.to_year
    if args.sort:
        config["sort"] = args.sort
    if args.mailto:
        config["mailto"] = args.mailto
    if args.out:
        config["out_excel"] = args.out

    if not config["query"]:
        print("请提供检索关键词：--query '...' 或在配置文件中设置 query")
        parser.print_help()
        return

    print("=" * 70)
    print("  Crossref 文献检索")
    print("=" * 70)
    print(f"  检索词     : {config['query']}")
    print(f"  最大数量   : {config['max_results']}")
    print(f"  排序方式   : {config['sort']}")
    if config.get("filter_from_year") or config.get("filter_to_year"):
        print(f"  年份范围   : {config.get('filter_from_year','')} ~ {config.get('filter_to_year','')}")
    if config.get("filter_type"):
        print(f"  文献类型   : {config['filter_type']}")
    if config.get("mailto"):
        print(f"  联系邮箱   : {config['mailto']}")

    client = CrossrefClient(
        mailto=config.get("mailto"),
        sleep_sec=config["sleep_sec"],
        timeout=config["timeout"]
    )

    print(f"\n开始检索...")
    raw_items = client.search_all(
        query=config["query"],
        max_results=config["max_results"],
        batch_size=config["batch_size"],
        sort=config["sort"],
        order=config["order"],
        filter_from_year=config.get("filter_from_year"),
        filter_to_year=config.get("filter_to_year"),
        filter_type=config.get("filter_type"),
        filter_has_abstract=config.get("filter_has_abstract", False),
    )

    print(f"共获取 {len(raw_items)} 篇文献")

    records = [parse_item(item) for item in raw_items]
    export_excel(records, config["out_excel"])
    export_json(records, raw_items, config["out_json"])

    print("\n" + "=" * 70)
    print("  ✅ 完成！")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    main()
