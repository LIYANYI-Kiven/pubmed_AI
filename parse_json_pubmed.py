#!/usr/bin/env python3
"""
parse_json_pubmed.py

整理 pubmed_search.py 生成的 JSON 文件，
提取每篇文章的：标题、PubMed ID、DOI、PubMed 摘要链接。

用法:
    python parse_json_pubmed.py
    python parse_json_pubmed.py --input output/pubmed_results.json --output output/pubmed_summary.xlsx
"""

import argparse
import json
import os
from typing import List, Dict

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


def parse_pubmed_json(input_file: str) -> List[Dict]:
    """从 pubmed_search.py 生成的 JSON 中提取核心字段。"""
    with open(input_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    articles = data.get("articles", [])
    records = []
    for i, article in enumerate(articles, 1):
        pmid = str(article.get("pmid") or "")
        doi  = str(article.get("doi")  or "")

        pubmed_url = article.get("links", {}).get("pubmed", "")
        if not pubmed_url and pmid:
            pubmed_url = f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"

        doi_url = article.get("links", {}).get("doi", "")
        if not doi_url and doi:
            doi_url = f"https://doi.org/{doi}"

        records.append({
            "No.":              i,
            "Title":            article.get("title", ""),
            "PubMed ID":        pmid,
            "DOI":              doi,
            "PubMed Abstract":  pubmed_url,
            "DOI URL":          doi_url,
            "Publication Date": article.get("publication_date", ""),
        })

    return records


def export_excel(records: List[Dict], out_file: str):
    df = pd.DataFrame(records).set_index("No.")
    os.makedirs(os.path.dirname(out_file) if os.path.dirname(out_file) else ".", exist_ok=True)

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Summary")
        ws = writer.sheets["Summary"]

        fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        col_widths = {"A": 6, "B": 70, "C": 12, "D": 35, "E": 45, "F": 45, "G": 14}
        for col, w in col_widths.items():
            ws.column_dimensions[col].width = w

        even = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
        odd  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.fill = even if row_idx % 2 == 0 else odd
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row_idx].height = 40

    print(f"  📊 Excel: {out_file}（{len(df)} 篇）")


def main():
    parser = argparse.ArgumentParser(description="整理 pubmed_search.py 生成的 JSON")
    parser.add_argument("--input",  default="output/pubmed_results.json",  help="输入 JSON 路径")
    parser.add_argument("--output", default="output/pubmed_summary.xlsx",  help="输出 Excel 路径")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"❌ 文件不存在: {args.input}")
        return

    print(f"读取: {args.input}")
    records = parse_pubmed_json(args.input)
    print(f"共 {len(records)} 篇文献")

    export_excel(records, args.output)

    # 同时输出一个精简 JSON
    out_json = os.path.splitext(args.output)[0] + "_detail.json"
    with open(out_json, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    print(f"  📄 JSON:  {out_json}")


if __name__ == "__main__":
    main()
