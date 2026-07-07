#!/usr/bin/env python3
"""
NCBI LitSense 2.0 语义检索工具

LitSense 2.0 是 NCBI 提供的生物医学语义搜索系统，支持跨 PubMed 摘要
和 PMC Open Access 全文的句子级与段落级语义检索。
覆盖约 38 million PubMed 摘要和 6.6 million PMC 全文。

API 端点：
  句子模式：https://www.ncbi.nlm.nih.gov/research/litsense2-api/api/sentences/
  段落模式：https://www.ncbi.nlm.nih.gov/research/litsense2-api/api/passages/

用法：
  python litsense_search.py --query "glycosylation sites in rice proteins" --mode sentence
  python litsense_search.py --query "plant protein N-glycosylation identified by mass spectrometry" --mode passage
  python litsense_search.py --config litsense_config.json
"""

import argparse
import json
import os
import time
from typing import List, Dict, Optional

import requests
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill


# ---------------------------------------------------------------------------
# LitSense 2.0 API 客户端
# ---------------------------------------------------------------------------

class LitSenseClient:
    """
    NCBI LitSense 2.0 API 客户端
    支持句子（sentence）和段落（passage）两种检索模式。

    API 说明：
      - 每次查询最多返回 100 条结果
      - 无需 API Key，免费开放
      - 支持自然语言查询，不需要布尔检索式
      - 适合用已知句子/段落进行语义相似度检索
    """

    SENTENCE_URL = "https://www.ncbi.nlm.nih.gov/research/litsense2-api/api/sentences/"
    PASSAGE_URL  = "https://www.ncbi.nlm.nih.gov/research/litsense2-api/api/passages/"

    SECTION_TYPES = [
        "abstract", "intro", "methods", "results",
        "discuss", "concl", "reference", "other"
    ]

    def __init__(self, sleep_sec: float = 1.0, timeout: float = 60.0):
        self.sleep_sec = sleep_sec
        self.timeout = timeout
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": "pubmed-litsense/1.0 (research tool)"
        })

    def search(
        self,
        query: str,
        mode: str = "sentence",
        rerank: bool = True,
        section: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
    ) -> List[Dict]:
        """
        执行语义检索。

        Args:
            query:     自然语言查询（句子或段落）
            mode:      检索模式，"sentence" 或 "passage"
            rerank:    是否使用 MedCPT 语义重排，建议开启
            section:   可选，限定文献章节类型
                       可选值: abstract/intro/methods/results/discuss/concl/reference/other
            date_from: 可选，发表日期起始（格式 YYYY/MM/DD）
            date_to:   可选，发表日期截止（格式 YYYY/MM/DD）

        Returns:
            结果列表，每条包含 pmid/pmcid、文本内容、来源、section 等字段
        """
        if mode == "sentence":
            url = self.SENTENCE_URL
        elif mode == "passage":
            url = self.PASSAGE_URL
        else:
            raise ValueError(f"mode 必须是 'sentence' 或 'passage'，当前值: {mode}")

        params: Dict = {
            "query": query,
            "rerank": "true" if rerank else "false",
        }
        if section:
            params["section"] = section
        if date_from:
            params["date_from"] = date_from
        if date_to:
            params["date_to"] = date_to

        try:
            r = self.session.get(url, params=params, timeout=self.timeout)
            r.raise_for_status()
            data = r.json()
            return data if isinstance(data, list) else data.get("results", [])
        except requests.exceptions.Timeout:
            print(f"  ⚠️  请求超时（{self.timeout}s）。LitSense 段落检索平均需 17s，句子检索约 9s。")
            raise
        except requests.exceptions.HTTPError as e:
            print(f"  ⚠️  HTTP 错误: {e}")
            raise

    def search_multiple(
        self,
        queries: List[str],
        mode: str = "sentence",
        rerank: bool = True,
        section: Optional[str] = None,
        date_from: Optional[str] = None,
        date_to: Optional[str] = None,
    ) -> Dict[str, List[Dict]]:
        """
        对多个查询依次执行检索，返回 {query: results} 字典。
        每次查询之间自动等待 sleep_sec 秒。
        """
        all_results = {}
        for i, query in enumerate(queries, 1):
            print(f"  [{i}/{len(queries)}] 检索: {query[:80]}...")
            try:
                results = self.search(
                    query=query, mode=mode, rerank=rerank,
                    section=section, date_from=date_from, date_to=date_to
                )
                all_results[query] = results
                print(f"    → 返回 {len(results)} 条结果")
            except Exception as e:
                print(f"    → 失败: {e}")
                all_results[query] = []

            if i < len(queries):
                time.sleep(self.sleep_sec)

        return all_results


# ---------------------------------------------------------------------------
# 结果解析与格式化
# ---------------------------------------------------------------------------

def parse_result(item: Dict, query: str, mode: str, rank: int) -> Dict:
    """将单条 API 结果规范化为统一字段。"""
    return {
        "Query":         query,
        "Mode":          mode,
        "Rank":          rank,
        "PMID":          item.get("pmid", ""),
        "PMCID":         item.get("pmcid", ""),
        "Title":         item.get("article_title", item.get("title", "")),
        "Authors":       _fmt_authors(item.get("authors", [])),
        "Journal":       item.get("journal", ""),
        "Year":          item.get("year", ""),
        "Section":       item.get("section", ""),
        "Text":          item.get("text", item.get("passage", item.get("sentence", ""))),
        "Score":         round(float(item.get("score", 0)), 4),
    }


def _fmt_authors(authors) -> str:
    if not authors:
        return ""
    if isinstance(authors, list):
        names = []
        for a in authors:
            if isinstance(a, dict):
                names.append(a.get("name", str(a)))
            else:
                names.append(str(a))
        return "; ".join(names[:5]) + (" et al." if len(names) > 5 else "")
    return str(authors)


def results_to_df(all_results: Dict[str, List[Dict]], mode: str) -> pd.DataFrame:
    """将所有查询结果合并为 DataFrame。"""
    rows = []
    for query, items in all_results.items():
        for rank, item in enumerate(items, 1):
            rows.append(parse_result(item, query, mode, rank))
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def export_excel(df: pd.DataFrame, out_file: str):
    """导出格式化 Excel。"""
    if df.empty:
        print("  ⚠️  无结果可导出")
        return

    os.makedirs(os.path.dirname(out_file) if os.path.dirname(out_file) else ".", exist_ok=True)

    with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="LitSense Results")
        ws = writer.sheets["LitSense Results"]

        header_fill = PatternFill(start_color="2E6B9E", end_color="2E6B9E", fill_type="solid")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        col_widths = {
            "A": 50, "B": 10, "C": 6, "D": 12, "E": 12,
            "F": 60, "G": 30, "H": 20, "I": 8, "J": 12, "K": 80, "L": 8
        }
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width

        even_fill = PatternFill(start_color="EAF2FB", end_color="EAF2FB", fill_type="solid")
        odd_fill  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.fill = even_fill if row_idx % 2 == 0 else odd_fill
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.row_dimensions[row_idx].height = 60

    print(f"  📁 已导出: {out_file}（{len(df)} 条结果）")


def export_json(all_results: Dict, out_file: str):
    """将原始结果保存为 JSON。"""
    os.makedirs(os.path.dirname(out_file) if os.path.dirname(out_file) else ".", exist_ok=True)
    with open(out_file, "w", encoding="utf-8") as f:
        json.dump(all_results, f, ensure_ascii=False, indent=2)
    print(f"  📁 原始结果: {out_file}")


# ---------------------------------------------------------------------------
# 配置加载
# ---------------------------------------------------------------------------

def load_config(path: str) -> Dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


DEFAULT_CONFIG = {
    "queries":    [],
    "mode":       "sentence",
    "rerank":     True,
    "section":    None,
    "date_from":  None,
    "date_to":    None,
    "sleep_sec":  1.5,
    "timeout":    90.0,
    "out_excel":  "output/litsense_results.xlsx",
    "out_json":   "output/litsense_raw.json",
}


# ---------------------------------------------------------------------------
# 主函数
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="NCBI LitSense 2.0 语义检索工具")
    parser.add_argument("--query",  type=str, help="单条查询语句（自然语言）")
    parser.add_argument("--mode",   type=str, default="sentence",
                        choices=["sentence", "passage"],
                        help="检索模式：sentence（句子）或 passage（段落）")
    parser.add_argument("--config", type=str, help="JSON 配置文件路径")
    parser.add_argument("--out",    type=str, help="输出 Excel 文件路径")
    args = parser.parse_args()

    # 加载配置
    config = dict(DEFAULT_CONFIG)
    if args.config and os.path.exists(args.config):
        config.update(load_config(args.config))

    # 命令行参数覆盖
    if args.query:
        config["queries"] = [args.query]
    if args.mode:
        config["mode"] = args.mode
    if args.out:
        config["out_excel"] = args.out

    if not config["queries"]:
        print("请提供查询语句：--query '...' 或在配置文件中设置 queries 列表")
        parser.print_help()
        return

    print("=" * 70)
    print(f"  NCBI LitSense 2.0 语义检索")
    print("=" * 70)
    print(f"  模式       : {config['mode']}")
    print(f"  查询数量   : {len(config['queries'])}")
    print(f"  语义重排   : {config['rerank']}")
    if config.get("section"):
        print(f"  章节过滤   : {config['section']}")
    if config.get("date_from") or config.get("date_to"):
        print(f"  日期范围   : {config.get('date_from','')} ~ {config.get('date_to','')}")

    client = LitSenseClient(
        sleep_sec=config["sleep_sec"],
        timeout=config["timeout"]
    )

    print(f"\n开始检索...")
    all_results = client.search_multiple(
        queries=config["queries"],
        mode=config["mode"],
        rerank=config["rerank"],
        section=config.get("section"),
        date_from=config.get("date_from"),
        date_to=config.get("date_to"),
    )

    # 统计
    total = sum(len(v) for v in all_results.values())
    print(f"\n共返回 {total} 条结果")

    # 导出
    df = results_to_df(all_results, config["mode"])
    export_excel(df, config["out_excel"])
    export_json(all_results, config["out_json"])

    print("\n" + "=" * 70)
    print("  ✅ 完成！")
    print("=" * 70 + "\n")


if __name__ == "__main__":
    main()
